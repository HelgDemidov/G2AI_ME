"""Встроенные чарты xlsx (spec convert-xlsx §3): архитектурный аналог
``docx_groups.py``, адаптированный под безпотоковую (cell-anchored) модель —
у xlsx-чарта нет «места в абзаце», есть лист + якорная ячейка, поэтому вместо
сентинел-в-потоке маркер декларативно позиционируется сразу после таблицы
своего листа (``converters._convert_xlsx``). Отсюда и другое имя функции
детекта: в отличие от docx-аналога (``extract_and_strip_groups`` — переписывает
``word/document.xml``, вырезая группу), здесь ``raw`` НЕ модифицируется вовсе —
называть это «strip» было бы неточно.

Детект: ``xl/worksheets/sheetK.xml`` несёт ``<drawing r:id="rIdX"/>`` ->
``xl/worksheets/_rels/sheetK.xml.rels`` -> ``xl/drawings/drawingM.xml`` ->
``<xdr:oneCellAnchor>``/``<xdr:twoCellAnchor>`` с ``<xdr:from>`` (якорная
ячейка: 0-indexed col/row) и ``graphicFrame``, несущим ``c:chart`` (ссылка на
``xl/charts/chartN.xml`` через ``xl/drawings/_rels/drawingM.xml.rels``).
Имя листа -> путь его XML-парта — через ``xl/workbook.xml`` (имя -> r:id) +
``xl/_rels/workbook.xml.rels`` (r:id -> Target); три разных источника
относительных путей (workbook/лист/drawing) требуют resolve относительно
директории КАЖДОГО конкретного source-парта (не хардкод «xl/», как у docx,
где все rels живут в одной ``word/``, см. ``_resolve_target``).

``id12`` — sha256 XML-СТРУКТУРЫ чарта (``etree.tostring`` парсенного
``xl/charts/chartN.xml``), НЕ байтов рендера: тот же принцип, что у
docx-групп (единственное отклонение, на котором дважды спотыкались тесты
docx — здесь зафиксировано в докстроке заранее). ``captions`` — из
``c:title`` чарт-парта: идентичная DrawingML-схема ``c:tx/c:rich/a:p/a:r/a:t``,
что у нативных docx-чартов (§2-ter convert-docx) — код адаптируется под
другие XML-пути, не копируется 1:1.
"""
from __future__ import annotations

import hashlib
import posixpath
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from lxml import etree
from openpyxl.utils import get_column_letter

_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
}
_NUMERIC_JUNK_RE = re.compile(r"^-?\d+$")  # posOffset/ext координаты в itertext(), см. docx_groups


def _q(prefix: str, local: str) -> str:
    return f"{{{_NS[prefix]}}}{local}"


@dataclass(frozen=True)
class XlsxChart:
    id12: str
    sheet: str
    anchor_cell: str
    captions: tuple[str, ...]


def _filter_caption_texts(texts: Any) -> tuple[str, ...]:
    """Тот же фильтр, что ``docx_groups._filter_caption_texts`` — отсеивает
    числовой геометрический мусор и дедуплицирует по порядку появления."""
    seen: set[str] = set()
    out: list[str] = []
    for t in texts:
        s = t.strip()
        if not s or _NUMERIC_JUNK_RE.match(s) or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return tuple(out)


def _rel_targets(z: zipfile.ZipFile, part: str) -> dict[str, str]:
    """rId -> сырой Target (ещё не resolve-нутый) для ``part``, из соседнего .rels."""
    rels_name = f"{posixpath.dirname(part)}/_rels/{posixpath.basename(part)}.rels"
    if rels_name not in z.namelist():
        return {}
    root = etree.fromstring(z.read(rels_name))
    return {rel.get("Id"): rel.get("Target") for rel in root if rel.get("Id")}


def _resolve_target(source_part: str, target: str) -> str:
    """OPC-резолв Target относительно СВОЕГО source-парта: абсолютный
    (ведущий ``/``) -> package-root, иначе -> относительно директории
    source_part (не хардкод одной директории — в отличие от docx, где все
    rels живут в ``word/``, здесь по цепочке участвуют ``xl/``,
    ``xl/worksheets/``, ``xl/drawings/``)."""
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))


def _sheet_parts(z: zipfile.ZipFile) -> dict[str, str]:
    """Имя листа (как в ``wb.sheetnames``) -> путь XML-парта листа в zip."""
    names = set(z.namelist())
    if "xl/workbook.xml" not in names:
        return {}
    rel_targets = _rel_targets(z, "xl/workbook.xml")
    root = etree.fromstring(z.read("xl/workbook.xml"))
    out: dict[str, str] = {}
    for sheet_el in root.findall(f".//{_q('main', 'sheet')}"):
        name, rid = sheet_el.get("name"), sheet_el.get(_q("r", "id"))
        if name is None or rid is None or rid not in rel_targets:
            continue
        part = _resolve_target("xl/workbook.xml", rel_targets[rid])
        if part in names:
            out[name] = part
    return out


def _chart_anchors(drawing_root: Any) -> list[tuple[Any, Any]]:
    """(anchor, chart_ref) для каждого ``xdr:oneCellAnchor``/``xdr:twoCellAnchor``,
    несущего ``c:chart`` где-то внутри (обычно в ``graphicFrame``)."""
    pairs: list[tuple[Any, Any]] = []
    for tag in ("oneCellAnchor", "twoCellAnchor"):
        for anchor in drawing_root.findall(_q("xdr", tag)):
            chart_ref = anchor.find(f".//{_q('c', 'chart')}")
            if chart_ref is not None:
                pairs.append((anchor, chart_ref))
    return pairs


def _anchor_col_row(anchor: Any) -> tuple[int, int]:
    """``xdr:from`` (0-indexed col/row) -> (col, row) 1-indexed, как
    ``openpyxl.utils.column_index_from_string``/``ws.cell`` ожидают."""
    frm = anchor.find(_q("xdr", "from"))
    return int(frm.findtext(_q("xdr", "col"))) + 1, int(frm.findtext(_q("xdr", "row"))) + 1


def _anchor_cell(anchor: Any) -> str:
    """``xdr:from`` (0-indexed col/row) -> Excel-ссылка на ячейку (``D2``)."""
    col, row = _anchor_col_row(anchor)
    return f"{get_column_letter(col)}{row}"


def _chart_title(chart_root: Any) -> tuple[str, ...]:
    title = chart_root.find(f".//{_q('c', 'title')}")
    if title is None:
        return ()
    return _filter_caption_texts(title.itertext())


def _iter_chart_entries(raw: Path) -> list[tuple[XlsxChart, Any]]:
    """Общий обход workbook, лист за листом, для ``extract_charts`` (метаданные)
    И ``extract_chart_roots`` (спек chart-data-extraction §4.1: data-driven
    рендер нужен УЖЕ распарсенный ``chart_root``, а не только его id12/captions) —
    один проход zip/XML на оба потребителя, а не два независимых. ``raw`` НЕ
    изменяется (см. докстроку модуля). Малформед/недостижимая ссылка на любом
    шаге цепочки (лист без drawing, drawing без rels, чарт-парт отсутствует) —
    честно пропускается (terminal safety net — конвертация не падает на
    повреждённом OOXML, симметрично ``_classify_docx``)."""
    with zipfile.ZipFile(raw) as z:
        names = set(z.namelist())
        entries: list[tuple[XlsxChart, Any]] = []
        for sheet_name, sheet_part in _sheet_parts(z).items():
            if sheet_part not in names:
                continue
            sheet_rels = _rel_targets(z, sheet_part)
            drawing_ref = etree.fromstring(z.read(sheet_part)).find(_q("main", "drawing"))
            if drawing_ref is None:
                continue
            drid = drawing_ref.get(_q("r", "id"))
            if drid is None or drid not in sheet_rels:
                continue
            drawing_part = _resolve_target(sheet_part, sheet_rels[drid])
            if drawing_part not in names:
                continue
            drawing_rels = _rel_targets(z, drawing_part)
            drawing_root = etree.fromstring(z.read(drawing_part))
            for anchor, chart_ref in _chart_anchors(drawing_root):
                crid = chart_ref.get(_q("r", "id"))
                if crid is None or crid not in drawing_rels:
                    continue
                chart_part = _resolve_target(drawing_part, drawing_rels[crid])
                if chart_part not in names:
                    continue
                chart_root = etree.fromstring(z.read(chart_part))
                entries.append(
                    (
                        XlsxChart(
                            id12=hashlib.sha256(etree.tostring(chart_root)).hexdigest()[:12],
                            sheet=sheet_name,
                            anchor_cell=_anchor_cell(anchor),
                            captions=_chart_title(chart_root),
                        ),
                        chart_root,
                    )
                )
        return entries


def extract_charts(raw: Path) -> list[XlsxChart]:
    """Все встроенные чарты workbook, лист за листом (метаданные — id12/sheet/
    anchor/captions, БЕЗ распарсенного XML; для data-driven рендера см.
    ``extract_chart_roots``)."""
    return [entry for entry, _root in _iter_chart_entries(raw)]


def extract_chart_roots(raw: Path) -> dict[str, Any]:
    """id12 -> распарсенный ``chart_root`` (spec chart-data-extraction §4.1) —
    вход для ``chart_data.parse_chart``. Разделено от ``extract_charts``
    (которому распарсенный XML не нужен снаружи), но обход общий
    (``_iter_chart_entries``) — воркбук не читается дважды."""
    return {entry.id12: root for entry, root in _iter_chart_entries(raw)}


def _chart_refs(chart_root: Any) -> list[tuple[str, str]]:
    """(sheet_name, cell_range_text) для каждой ``<c:f>`` формулы серии чарта
    (раскавычивание листа с пробелами: ``'My Sheet'!$A$1`` -> ``My Sheet``,
    ``''``->``'`` — стандартное Excel-экранирование апострофа)."""
    out: list[tuple[str, str]] = []
    for f in chart_root.findall(f".//{_q('c', 'f')}"):
        text = f.text or ""
        if "!" not in text:
            continue
        sheet, _, cell_range = text.partition("!")
        sheet = sheet.strip()
        if sheet.startswith("'") and sheet.endswith("'"):
            sheet = sheet[1:-1].replace("''", "'")
        if sheet:
            out.append((sheet, cell_range))
    return out


def render_chart_marker(chart: XlsxChart) -> str:
    caption_line = "; ".join(chart.captions) if chart.captions else "(нет текста)"
    return (
        f"> [Figure, xlsx chart {chart.id12} on {chart.sheet}!{chart.anchor_cell} — "
        f"chart content not analyzed]\n"
        f"> captions: {caption_line}"
    )
