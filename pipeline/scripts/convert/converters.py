"""Реестр конвертеров raw.* -> doc.md (Strategy). Ключ — расширение raw-файла.

Оркестратор (_do_convert) не знает форматов: resolve_converter(raw) -> Converter.
Новый формат = новая запись в _CONVERTERS (+ свой convert-модуль), ноль правок
оркестратора (чартер convert/architecture.md §3.1).
"""
from __future__ import annotations

import datetime as _dt
import hashlib
import io
import logging
import os
import posixpath
import re
import shutil
import subprocess
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Protocol
from xml.etree import ElementTree

import pdfplumber

from convert import cloud_ocr, html_preprocess, ocr_headings
from convert.pdf_to_markdown import convert as pdf_convert
from core import fsio, schema
from core.env import load_dotenv

logger = logging.getLogger(__name__)

_HTML_HEADING_RE = re.compile(rb"<h[1-6][\s>]", re.I)  # B2: детект <hN> в исходном HTML


class ConversionError(RuntimeError):
    """Базовый типизированный отказ конвертации (per-doc изоляция ловит как обычно)."""


class UnsupportedFormat(ConversionError):
    """Расширение raw.* не имеет зарегистрированного конвертера."""


class NeedsOCR(ConversionError):
    """PDF без текстового слоя — нужен OCR-путь (спек convert-ocr)."""


class ConvertFn(Protocol):
    """(raw, out, language, *, record) — ``record`` нужен облачным веткам (pdf):
    sensitivity-гейт (spec convert-cloud-tier §6). HTML/будущие не-облачные
    конвертеры параметр игнорируют — сигнатура едина для всех записей реестра."""

    def __call__(
        self, raw: Path, out: Path, language: str | None, *, record: schema.SourceRecord | None = None
    ) -> None: ...


@dataclass(frozen=True)
class Converter:
    name: str      # стабильный id пути конвертации ("pdf")
    version: str   # бамп => авто-реконверсия всех документов формата (needed_stages)
    convert: ConvertFn


SCAN_MIN_CHARS_PER_PAGE = 50      # страница «с текстом», если извлечено >= стольких символов
SCAN_MIN_TEXTPAGE_FRACTION = 0.5  # доля страниц с текстом ниже порога => скан


def _detect_scan(raw: Path) -> None:
    """Поднять NeedsOCR, если у большинства страниц нет текст-слоя.

    Отдельный дешёвый проход pdfplumber ДО pdf_convert: полный конвертер на скане
    дал бы пустой вывод с враньём «пустой файл» — диагноз должен называть причину
    (агенда §4: «явный флаг „нужен OCR“»). Двойной парс PDF — секунды, конвертация
    редка (раз на документ).
    """
    with pdfplumber.open(raw) as pdf:
        n = len(pdf.pages)
        if n == 0:
            return  # пустой PDF диагностирует pdf_convert («PDF без страниц»)
        with_text = sum(
            1 for p in pdf.pages
            if len((p.extract_text() or "").strip()) >= SCAN_MIN_CHARS_PER_PAGE
        )
    if with_text / n < SCAN_MIN_TEXTPAGE_FRACTION:
        raise NeedsOCR(
            f"{raw.name}: текст-слой лишь на {with_text}/{n} страниц — вероятен скан; "
            f"OCR-путь не реализован (см. docs/pipeline/convert/tech_specs/convert-ocr)"
        )


def _was_ocr_normalized(raw: Path) -> bool:
    """PDF уже прошёл OCR-нормализацию РАНЬШЕ (по метаданным ocrmypdf).

    `_ocr_normalize` мутирует `raw` in-place (один файл, не сайдкар) — после первого
    успеха текст-слой уже есть, и `_detect_scan` больше НЕ поднимет `NeedsOCR` на
    повторных конвертациях (`--force`, бамп версии конвертера). Без этой проверки
    `ocr_headings` перестал бы применяться после первого прогона — метаданные ocrmypdf
    (`Creator: ocrmypdf ...`) переживают мутацию текст-слоя и остаются надёжным маркером.
    """
    with pdfplumber.open(raw) as pdf:
        creator = (pdf.metadata.get("Creator") or "").lower()
    return "ocrmypdf" in creator


# rec.language (schema.py: ISO 639-1, либо 639-3 где нет 639-1, напр. cnr) -> tesseract langcode.
TESSERACT_LANGS = {
    "en": "eng", "et": "est", "sr": "srp_latn", "cnr": "srp_latn",
    "hr": "hrv", "bs": "bos", "sl": "slv", "sq": "sqi", "mk": "mkd",
    "de": "deu", "fr": "fra", "it": "ita", "es": "spa",
    "ru": "rus", "ar": "ara", "zh": "chi_sim", "ja": "jpn",
    # zh по умолчанию упрощённый (материк); традиционный (HK/TW) — chi_tra, добавить при нужде.
}
# CJK/арабский — БЕЗ +eng: удваивает проход и иногда интерферирует (иные скрипты).
_NO_ENG_SUFFIX = frozenset({"chi_sim", "chi_tra", "jpn", "ara"})


def _tesseract_langs(language: str | None) -> str:
    """rec.language -> tesseract -l аргумент. Латиница получает +eng (гос-документы часто
    со вставками EN); CJK/арабский — нет (см. _NO_ENG_SUFFIX). Неизвестный код -> честный
    eng-fallback с предупреждением (не молчаливая порча качества)."""
    mapped = TESSERACT_LANGS.get(language or "en")
    if mapped is None:
        logger.warning("неизвестный языковой код %r для OCR — используется eng", language)
        mapped = "eng"
    if mapped == "eng" or mapped in _NO_ENG_SUFFIX:
        return mapped
    return f"{mapped}+eng"


OCR_TIMEOUT = 7200      # 2 ч — потолок для ~200-страничного скана на i5-6200U
OCR_PAGE_WARN = 200     # страниц > порога -> лог оценки времени до запуска
_OCR_STDERR_TAIL = 500  # символов stderr в ConversionError — достаточно для диагноза


def _check_langs_available(langs: str) -> None:
    """tesseract --list-langs -> ConversionError с apt-командой, если traineddata нет.

    Проверяется ДО (потенциально долгого) ocrmypdf — быстрый честный отказ вместо
    невнятной ошибки из недр ocrmypdf/tesseract.
    """
    result = subprocess.run(
        ["tesseract", "--list-langs"], check=False, capture_output=True, text=True
    )
    installed = set(result.stdout.splitlines()[1:])  # первая строка — заголовок списка
    missing = [code for code in langs.split("+") if code not in installed]
    if missing:
        apt_pkgs = " ".join(f"tesseract-ocr-{code.replace('_', '-')}" for code in missing)
        raise ConversionError(f"нет traineddata для {', '.join(missing)} — sudo apt install {apt_pkgs}")


def _capture_original_sha256(raw: Path) -> None:
    """Sha256 raw ДО OCR-мутации (spec ocr-eval-harness §8.1, S1) — `sha256` в
    `.state.yaml` отражает файл ПОСЛЕ ocrmypdf, издательский оригинал иначе не
    восстановим. Тот же паттерн, что `_cached_or_call_cloud` (read-modify-write
    `.state.yaml` напрямую по `raw.parent`, без протаскивания через `ConvertFn` —
    его сигнатура едина для всех форматов реестра, менять её ради одного поля
    OCR-ветки не стоит). Пишет ТОЛЬКО если поле ещё `None`: повторная нормализация
    (`--force`/бамп версии конвертера) уже мутированного raw не должна затереть
    РАНЕЕ захваченный оригинальный хэш пересчитанным от файла, который сам уже
    не оригинал."""
    state_path = raw.parent / ".state.yaml"
    state = schema.load_state(state_path)
    if state.original_sha256 is not None:
        return
    state.original_sha256 = fsio.sha256_file(raw)
    schema.save_state(state_path, state)


def _ocr_normalize(raw: Path, language: str | None) -> None:
    """OCR-нормализовать скан IN-PLACE: `raw` заменяется версией с невидимым текст-слоем.

    Один PDF-файл на документ, без сайдкара `.ocr.pdf` (раньше кэш жил отдельным файлом —
    двойное хранение того же документа; убрано по решению пользователя). Кэширование
    получается «бесплатно» иначе: после успеха `raw` САМ содержит текст-слой, поэтому
    следующий `_detect_scan(raw)` больше не поднимет `NeedsOCR`, и `_ocr_normalize` не
    вызовется повторно без явного `--force`/бампа версии конвертера. Вызывающий
    (`_do_convert` в run_pipeline.py) ОБЯЗАН пересчитать sha256/размер/mtime в
    `.state.yaml` после конвертации — raw физически изменился (его собственный
    `load_state`/`save_state`-раунд-трип идёт ПОСЛЕ этой функции и подхватит уже
    записанный `_capture_original_sha256` результат с диска — гонки нет).
    """
    if shutil.which("ocrmypdf") is None:
        raise NeedsOCR(
            f"{raw.name}: ocrmypdf не установлен — sudo apt install ocrmypdf "
            f"tesseract-ocr-srp-latn tesseract-ocr-est"
        )

    langs = _tesseract_langs(language)
    _check_langs_available(langs)

    with pdfplumber.open(raw) as pdf:
        n = len(pdf.pages)
    if n > OCR_PAGE_WARN:
        logger.warning(
            "%s: %d страниц — OCR займёт ориентировочно %d–%d мин",
            raw.name, n, n * 20 // 60, n * 40 // 60,
        )

    _capture_original_sha256(raw)  # ДО staging/subprocess — raw ещё не мутирован
    staging = fsio.staging_path(raw)
    result = subprocess.run(
        [
            "ocrmypdf", "--skip-text", "-l", langs, "--output-type", "pdf", "--quiet",
            str(raw), str(staging),
        ],
        check=False, capture_output=True, text=True, timeout=OCR_TIMEOUT,
    )
    if result.returncode != 0:
        raise ConversionError(
            f"{raw.name}: ocrmypdf завершился с кодом {result.returncode}: "
            f"{result.stderr[-_OCR_STDERR_TAIL:]}"
        )
    staging.replace(raw)


_CLOUD_DISABLED = False   # spec convert-cloud-tier §6.3 — settable через set_cloud_disabled (--no-cloud)
_CLOUD_KEY_WARNED = False  # warning про отсутствующий ключ — один раз за прогон, не на документ


def set_cloud_disabled(disabled: bool) -> None:
    """Единая точка для ``--no-cloud`` (run_pipeline.main()) — полностью отключает
    scan-OCR/figures-VLM пути, поведение = статус-кво до convert-cloud-tier."""
    global _CLOUD_DISABLED
    _CLOUD_DISABLED = disabled


def cloud_allowed(record: schema.SourceRecord | None) -> bool:
    """Гейты §6 спека convert-cloud-tier — единый предикат для scan-OCR и (позже)
    figures-VLM: --no-cloud -> sensitivity (ЛАТЕНТНЫЙ режим, один предикат) ->
    ключ. Порядок — от дешёвой проверки к дорогой (файл .env)."""
    global _CLOUD_KEY_WARNED
    if _CLOUD_DISABLED:
        return False
    if record is not None and record.sensitivity is schema.Sensitivity.confidential:
        return False
    load_dotenv()
    if not os.environ.get("OPENROUTER_API_KEY"):
        if not _CLOUD_KEY_WARNED:
            logger.warning("нет OPENROUTER_API_KEY — облачный OCR/figures отключены, локальный путь")
            _CLOUD_KEY_WARNED = True
        return False
    return True


def _cached_or_call_cloud(raw: Path, language: str | None, *, model: str) -> str | None:
    """Кэш-реконсиляция §2.2: валиден (файл существует ∧ sha256 ∧ модель совпадают)
    -> сеть не трогается. Модель разошлась, но raw тот же -> авто-перевызов
    ЗАПРЕЩЁН (сюрприз-биллинг) — используем кэш, сигнализируем warning'ом.
    Иначе (сайдкара нет ИЛИ raw изменился) -> облачный вызов; отказ после
    ретраев -> None + warning (локальный фолбэк, документ не падает)."""
    cache_path = cloud_ocr.cache_path(raw)
    state_path = raw.parent / ".state.yaml"
    state = schema.load_state(state_path)
    raw_sha256 = fsio.sha256_file(raw)

    if cache_path.exists() and state.cloud_ocr_raw_sha256 == raw_sha256:
        if state.cloud_ocr_model != model:
            logger.warning(
                "%s: .cloudocr.md от модели %r, активна %r — используется кэш; "
                "для перевызова удалите .cloudocr.md",
                raw.name, state.cloud_ocr_model, model,
            )
        return cache_path.read_text(encoding="utf-8")

    try:
        text = cloud_ocr.convert_scan(raw, language, model=model)
    except Exception as exc:  # noqa: BLE001 — отказ облака после ретраев -> локальный фолбэк, не крах
        logger.warning("%s: облачный OCR не удался (%s) — локальный путь", raw.name, exc)
        return None

    fsio.atomic_write_text(cache_path, text)
    state.cloud_ocr_model = model
    state.cloud_ocr_raw_sha256 = raw_sha256
    schema.save_state(state_path, state)
    return text


def _convert_pdf(
    raw: Path, out: Path, language: str | None, *, record: schema.SourceRecord | None = None
) -> None:
    try:
        _detect_scan(raw)
        scanned = _was_ocr_normalized(raw)  # текст есть — но, может, уже был нормализован раньше
    except NeedsOCR:
        _ocr_normalize(raw, language)   # мутирует raw IN-PLACE (текст-слой встроен) — witness ВСЕГДА
        scanned = True
    if scanned and cloud_allowed(record):
        text = _cached_or_call_cloud(raw, language, model=cloud_ocr.ACTIVE_MODEL)
        if text is not None:
            # Additive-режим (§2.5, v2.1): облачная иерархия — главный производитель,
            # прецизионные тиры лишь ДОБАВЛЯЮТ пропущенное облаком (главы I.–VIII. me-crps);
            # полный promote_flat_headings (снятие+переоценка) уничтожил бы её.
            out.write_text(ocr_headings.merge_missing_headings(text), encoding="utf-8")
            return
    pdf_convert(str(raw), str(out))  # существующий конвертер, без изменений (локальный путь)
    if scanned:  # только OCR-ветка: цифровой путь не трогаем (размер-кластеризация там чище)
        out.write_text(
            ocr_headings.promote_flat_headings(out.read_text(encoding="utf-8")),
            encoding="utf-8",
        )


def _convert_html(
    raw: Path, out: Path, language: str | None, *, record: schema.SourceRecord | None = None
) -> None:
    import trafilatura  # ленивый импорт: pdf-путь не платит за html-зависимость

    raw_bytes = raw.read_bytes()
    html = html_preprocess.apply(raw_bytes)  # первый сматчившийся препроцессор (ELI и т.д.), иначе no-op
    text = trafilatura.extract(
        html,                            # bytes: charset определяет trafilatura
        output_format="markdown",
        include_tables=True,
        include_links=False,            # URL-хвосты — шум для чанков/эмбеддера
        include_images=False,
        favor_recall=True,              # гос-страницы: лучше лишний блок, чем потерянная статья
        with_metadata=False,            # frontmatter — производная meta.yaml, не trafilatura
    )
    if not text or not text.strip():
        raise ConversionError(f"{raw.name}: trafilatura не извлекла контента")
    if _HTML_HEADING_RE.search(raw_bytes) and not any(line.startswith("#") for line in text.splitlines()):
        # B2: генерическая ловушка trafilatura «главный контент без <article> теряет <hN>»
        # (эмпирика convert-html) — не отказ (favor recall: текст выжил), сигнал в лог.
        logger.warning(
            "%s: исходный HTML содержит <hN>, но в выходе ни одного markdown-заголовка — "
            "вероятна потеря структуры, кандидат на препроцессор (convert/html_preprocess.py)",
            raw.name,
        )
    out.write_text(text, encoding="utf-8")


DOCX_IMAGE_MIN_BYTES = 5000
# Калибровка на живой фикстуре (tests/fixtures/local/, 2026-07-19, 167 файлов
# word/media/*): чёткий естественный разрыв — 101 файл < 5000 байт (кластер
# мелких иконок/флагов легенды, медиана по всему набору — 2012 байт), дальше
# редкие крупные диаграммы (17 файлов > 100 КБ). Начальное значение, как и все
# численные пороги проекта — подлежит пересмотру по факту живой приёмки.


_DOCX_MARKER_SRC_PREFIX = "docx-marker:"


def _docx_referenced_media_ids(raw: Path) -> frozenset[str]:
    """id12-множество media-файлов, реально референсированных хоть одной XML-частью
    документа (document.xml/headers/footers/notes/charts — любая часть с собственным
    .rels). Файл в word/media/ БЕЗ единой ссылки — сирота: Word его не отображает,
    это не контент документа (реальный кейс — тестовая вырезка: слайсинг оставил
    media/rels полного отчёта, 21 из 28 больших изображений оказались сиротами,
    включая 4 фото людей; маркер для сироты — мусор в doc.md и лишний VLM-расход).

    Детекция ссылки — подстрока '"rIdN"' в байтах части: кавычки исключают
    префикс-коллизию (rId3 не матчится внутри rId30); недобор невозможен, возможный
    перебор (rId в видимом тексте) безопасен — лишний честный маркер, ровно
    поведение v1."""
    referenced: set[str] = set()
    with zipfile.ZipFile(raw) as z:
        names = set(z.namelist())
        for part in sorted(names):
            if not (part.startswith("word/") and part.endswith(".xml")) or "/_rels/" in part:
                continue
            rels_name = f"{posixpath.dirname(part)}/_rels/{posixpath.basename(part)}.rels"
            if rels_name not in names:
                continue
            try:
                rels_root = ElementTree.fromstring(z.read(rels_name))
            except ElementTree.ParseError:
                continue
            part_bytes = z.read(part)
            for rel in rels_root:
                if not rel.get("Type", "").endswith("/image"):
                    continue
                rid = rel.get("Id")
                target = rel.get("Target", "")
                if not rid or f'"{rid}"'.encode() not in part_bytes:
                    continue
                media = posixpath.normpath(posixpath.join(posixpath.dirname(part), target))
                if media.startswith("word/media/") and media in names:
                    referenced.add(hashlib.sha256(z.read(media)).hexdigest()[:12])
    return frozenset(referenced)


def _docx_image_markers(raw: Path, *, placed: frozenset[str] = frozenset()) -> str:
    """§2-bis фолбэк-проход (v2): маркеры под ``## Figures (position unknown)``
    ТОЛЬКО для media, которые (а) реально референсированы документом
    (``_docx_referenced_media_ids`` — orphan-фильтр), (б) не мельче
    ``DOCX_IMAGE_MIN_BYTES``, (в) не размещены инлайн mammoth-проходом
    (``placed``). Дедуп по id12: одинаковые байты под двумя именами
    (jpg/jpeg-двойники Word) дают один маркер. Пустой результат ("") валиден
    и ОЖИДАЕМ: на чистом документе весь реальный растр ложится инлайн."""
    referenced = _docx_referenced_media_ids(raw)
    lines: list[str] = []
    seen: set[str] = set()
    with zipfile.ZipFile(raw) as z:
        for name in sorted(z.namelist()):
            if not name.startswith("word/media/"):
                continue
            data = z.read(name)
            if len(data) < DOCX_IMAGE_MIN_BYTES:
                continue
            id12 = hashlib.sha256(data).hexdigest()[:12]
            if id12 in placed or id12 in seen or id12 not in referenced:
                continue
            seen.add(id12)
            lines.append(f"> [Image, docx media {id12} — raster content not analyzed]")
    if not lines:
        return ""
    return "\n## Figures (position unknown)\n\n" + "\n".join(lines) + "\n"


def _convert_docx(
    raw: Path, out: Path, language: str | None, *, record: schema.SourceRecord | None = None
) -> None:
    """v3 (§2-bis.2/§2-bis.3/§2-ter): прямой mammoth + свой markdownify-сабкласс,
    без markitdown. Пре-проход ``docx_groups`` вырезает composite-группы
    (``mc:AlternateContent``/``wpg:wgp`` — Word рисует сложную инфографику
    группой фигур, mammoth обходит её поэлементно и распадает ОДНУ диаграмму
    на россыпь фрагментов, живой кейс §2-ter.1) ДО mammoth, заменяя каждую на
    текстовый сентинел; custom image-handler инлайнит маркер РОВНО в месте
    вхождения ОДИНОЧНОЙ картинки (вне групп); пост-проход заменяет сентинелы
    честным маркером группы с сохранёнными подписями (zero-loss без VLM).
    ``_docx_image_markers`` — страховка для referenced-but-not-walked случаев
    (класс: картинка только в mc:Choice при пустом mc:Fallback), media
    поглощённых групп исключены через ``placed`` — сироты по-прежнему
    отфильтрованы."""
    import mammoth  # ленивые импорты: pdf/html-пути не платят за docx-зависимости
    from markdownify import ATX, MarkdownConverter
    from mammoth import html as mammoth_html

    from convert import docx_groups

    rewritten, groups = docx_groups.extract_and_strip_groups(raw)

    class _DocxMarkdownify(MarkdownConverter):
        def convert_img(self, el: Any, text: str, parent_tags: Any) -> str:
            # Единственный источник <img> в этом HTML — convert_image ниже,
            # поэтому src всегда несёт префикс-сентинел.
            id12 = (el.attrs.get("src") or "")[len(_DOCX_MARKER_SRC_PREFIX) :]
            return f"\n\n> [Image, docx media {id12} — raster content not analyzed]\n\n"

    placed_ids: set[str] = set()

    def convert_image(image: Any) -> list[Any]:
        with image.open() as f:
            data = f.read()
        if len(data) < DOCX_IMAGE_MIN_BYTES:
            return []
        id12 = hashlib.sha256(data).hexdigest()[:12]
        placed_ids.add(id12)
        return [mammoth_html.element("img", {"src": f"{_DOCX_MARKER_SRC_PREFIX}{id12}"})]

    converted = mammoth.convert_to_html(io.BytesIO(rewritten), convert_image=convert_image)

    text = _DocxMarkdownify(heading_style=ATX).convert(converted.value).strip()
    if not text:
        raise ConversionError(f"{raw.name}: mammoth/markdownify не извлекли контента")
    text = docx_groups.inject_group_markers(text, groups)
    fallback = _docx_image_markers(raw, placed=frozenset(placed_ids) | docx_groups.all_media_ids(groups))
    out.write_text(text + "\n" + fallback, encoding="utf-8")


def _xlsx_cell_str(value: Any) -> str:
    """Значение ячейки -> текст ячейки GFM-таблицы (spec convert-xlsx §2).

    Даты/datetime (openpyxl авто-конвертирует дата-форматированные числовые
    ячейки, без доп. кода) -> ISO; целый float -> без ``.0``; остальное ->
    ``str()``. Переносы строк -> пробел (то же соглашение, что
    ``pdf_to_markdown.render_tables``: pipe-символы НЕ экранируются, тот же
    принятый и уже задокументированный компромисс, не новая проблема).
    Проценты/валюта/прочие ``number_format`` НЕ интерпретируются — сырое
    значение (Design rationale: «сомнение ⇒ сырое число»)."""
    if value is None:
        return ""
    if isinstance(value, (_dt.datetime, _dt.date)):
        text = value.isoformat()
    elif isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    return text.strip().replace("\n", " ")


def _sheet_is_empty(ws: Any) -> bool:
    return all(cell.value is None for row in ws.iter_rows() for cell in row)


def _sheet_table(ws: Any) -> str:
    """Лист -> GFM-таблица: первая строка используемого диапазона —
    структурный заголовок (не семантический — зеркалит уже принятое и
    проверенное на реальных документах поведение markdownify-таблиц docx без
    ``<thead>``, см. Design rationale), разделитель, остальные строки —
    данные. Слитые диапазоны: openpyxl в НЕ-read_only режиме уже отдаёт
    значение только якорной (верхней левой) ячейке — прочие ячейки диапазона
    ``MergedCell`` с ``value=None`` (эмпирически подтверждено, без
    дополнительного кода по ``merged_cells.ranges``)."""
    rows = [[_xlsx_cell_str(c.value) for c in row] for row in ws.iter_rows()]
    header, *body = rows
    lines = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * len(header)) + " |"]
    lines.extend("| " + " | ".join(row) + " |" for row in body)
    return "\n".join(lines)


def _render_xlsx_chart_block(chart: Any, chart_root: Any) -> str:
    """Data-driven рендер одного xlsx-чарта (spec chart-data-extraction §4.1/
    §4.4): ``parse_chart``->``render_chart``; пустое извлечение (нет numCache) ->
    caption-фолбэк (честный маркер, как раньше). Провенанс-строка (лист+якорь) —
    добавляется ЗДЕСЬ (контейнером), не в ``render_chart`` (container-agnostic):
    сохраняет позиционную привязку, которую нёс прежний VLM-маркер
    (``on {sheet}!{anchor}``), критично для retrieval по листу/якорю."""
    from convert import chart_data, chart_render, xlsx_charts

    rendered = chart_render.render_chart(chart_data.parse_chart(chart_root))
    if rendered is None:
        return xlsx_charts.render_chart_marker(chart)
    provenance = f"> лист {chart.sheet}, якорь {chart.anchor_cell}"
    return f"{provenance}\n\n{rendered}"


def _convert_xlsx(
    raw: Path, out: Path, language: str | None, *, record: schema.SourceRecord | None = None
) -> None:
    """v2 (spec convert-xlsx §2/§3 + chart-data-extraction §4.1): один лист =
    одна GFM-таблица под заголовком ``## {sheet}``, в порядке ``wb.sheetnames``
    (workbook, не алфавитный). Скрытые листы остаются частью документа с
    суффиксом «(hidden)» — НЕ orphan-фильтр docx (лист — часть документа
    автора, просто визуально свёрнута, см. Design rationale, честность vs
    privacy-фильтрация). Пустой лист -> честный маркер, без пустой таблицы —
    КРОМЕ случая, когда на нём всё же висит чарт без единой ячейки данных
    (chart-only лист): тогда маркер листа не эмитится, а чарт-блок(и) не
    теряются молча. Чарты данного листа — сразу после его таблицы, в
    детерминированном порядке по якорю (топ-лефт первым); каждый —
    data-driven (таблица+mermaid) с caption-фолбэком на пустое извлечение."""
    import openpyxl
    from openpyxl.utils.cell import coordinate_to_tuple

    from convert import xlsx_charts

    wb = openpyxl.load_workbook(raw, data_only=True, read_only=False)
    chart_roots = xlsx_charts.extract_chart_roots(raw)
    charts_by_sheet: dict[str, list[xlsx_charts.XlsxChart]] = {}
    for chart in xlsx_charts.extract_charts(raw):
        charts_by_sheet.setdefault(chart.sheet, []).append(chart)
    for charts in charts_by_sheet.values():
        charts.sort(key=lambda c: coordinate_to_tuple(c.anchor_cell))

    sections: list[str] = []
    any_content = False
    for name in wb.sheetnames:
        ws = wb[name]
        heading = f"## {name}" if ws.sheet_state == "visible" else f"## {name} (hidden)"
        sheet_charts = charts_by_sheet.get(name, [])
        if _sheet_is_empty(ws) and not sheet_charts:
            sections.append(f'{heading}\n\n> [Sheet "{name}" — empty, skipped]')
            continue
        any_content = True
        parts = [heading]
        if not _sheet_is_empty(ws):
            parts.append(_sheet_table(ws))
        if sheet_charts:
            parts.append(
                "\n\n".join(_render_xlsx_chart_block(c, chart_roots[c.id12]) for c in sheet_charts)
            )
        sections.append("\n\n".join(parts))
    if not any_content:
        raise ConversionError(f"{raw.name}: ни один лист workbook не содержит данных")
    out.write_text("\n\n".join(sections) + "\n", encoding="utf-8")


_CONVERTERS: dict[str, Converter] = {
    "pdf": Converter("pdf", "5", _convert_pdf),  # v5: raster region-id (convert-cloud-tier §4)
    "html": Converter("html", "1", _convert_html),
    "docx": Converter("docx", "3", _convert_docx),  # v3: composite-группы (§2-ter)
    "xlsx": Converter("xlsx", "2", _convert_xlsx),  # v2: data-driven чарты (chart-data-extraction §4.1)
}


def resolve_converter(raw: Path) -> Converter:
    ext = raw.suffix.lstrip(".").lower()
    conv = _CONVERTERS.get(ext)
    if conv is None:
        known = ", ".join(sorted(_CONVERTERS))
        raise UnsupportedFormat(f"{raw.name}: формат '{ext}' не поддержан (есть: {known})")
    return conv
