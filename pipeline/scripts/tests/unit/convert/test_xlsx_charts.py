"""Тесты xlsx_charts.py (spec convert-xlsx §3 + chart-data-extraction §4.3):
детект встроенных чартов, captions из c:title, id12 по XML-структуре чарта.
Ни сети, ни LibreOffice — чистый XML in-memory (openpyxl.chart строит
реальный chart-парт). Мутуал-ownership/рендер-изоляция (``extract_chart_workbook``
и вся геометрия печатной области) удалены вместе с VLM-путём — см.
``chart_data.py``/``chart_render.py`` за data-driven заменой."""
from __future__ import annotations

import io
import zipfile
from pathlib import Path
from typing import Any

from lxml import etree
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from convert.xlsx_charts import (
    _chart_refs,
    _filter_caption_texts,
    _q,
    _rel_targets,
    _resolve_target,
    _sheet_parts,
    extract_charts,
)


def _active(wb: Any) -> Any:
    """``wb.active`` типизирован как ``Worksheet | Chartsheet | None`` в
    стабах — свежий ``Workbook()`` всегда несёт активный лист-Worksheet."""
    ws = wb.active
    assert ws is not None
    return ws


def _workbook_with_chart(
    tmp_path: Path,
    *,
    title: str | None = "Chart Title",
    anchor: str = "D2",
    sheet_name: str = "Data",
    file_name: str = "raw.xlsx",
) -> Path:
    wb = Workbook()
    ws = _active(wb)
    ws.title = sheet_name
    ws.append(["Cat", "Val"])
    ws.append(["A", 1])
    ws.append(["B", 2])
    chart = BarChart()
    if title is not None:
        chart.title = title
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)
    raw = tmp_path / file_name
    wb.save(raw)
    return raw


def test_extract_charts_empty_workbook_returns_empty_list(tmp_path: Path) -> None:
    wb = Workbook()
    raw = tmp_path / "raw.xlsx"
    wb.save(raw)
    assert extract_charts(raw) == []


def test_extract_charts_detects_single_chart_with_title_and_anchor(tmp_path: Path) -> None:
    raw = _workbook_with_chart(tmp_path, title="Costs of LTE and 5G", anchor="D2", sheet_name="Data")
    charts = extract_charts(raw)
    assert len(charts) == 1
    chart = charts[0]
    assert chart.sheet == "Data"
    assert chart.anchor_cell == "D2"
    assert chart.captions == ("Costs of LTE and 5G",)


def test_extract_charts_no_title_gives_empty_captions(tmp_path: Path) -> None:
    raw = _workbook_with_chart(tmp_path, title=None)
    charts = extract_charts(raw)
    assert len(charts) == 1
    assert charts[0].captions == ()


def test_extract_charts_id12_stable_across_repeated_calls(tmp_path: Path) -> None:
    raw = _workbook_with_chart(tmp_path)
    id1 = extract_charts(raw)[0].id12
    id2 = extract_charts(raw)[0].id12
    assert id1 == id2
    assert len(id1) == 12


def test_extract_charts_different_titles_get_distinct_ids(tmp_path: Path) -> None:
    """Билдер рассчитан на один чарт на документ (симметрично docx-тестам) —
    сверяем на двух отдельных однокграфиковых документах, что id12 зависит от
    содержимого чарта (заголовок входит в chart-парт)."""
    raw_a = _workbook_with_chart(tmp_path, title="Title A", file_name="a.xlsx")
    raw_b = _workbook_with_chart(tmp_path, title="Title B", file_name="b.xlsx")
    id_a = extract_charts(raw_a)[0].id12
    id_b = extract_charts(raw_b)[0].id12
    assert id_a != id_b


def test_extract_charts_anchor_cell_reflects_position(tmp_path: Path) -> None:
    raw = _workbook_with_chart(tmp_path, anchor="G10")
    charts = extract_charts(raw)
    assert charts[0].anchor_cell == "G10"


# --- _filter_caption_texts: числовой мусор / дедуп / пустые строки (чистая функция) ---


def test_filter_caption_texts_skips_numeric_junk_duplicates_and_blanks() -> None:
    assert _filter_caption_texts(["Title", "42", "-3", "Title", "  ", "Subtitle"]) == (
        "Title",
        "Subtitle",
    )


# --- _rel_targets / _resolve_target (чистые функции, малформед-цепочка) ---


def test_rel_targets_missing_rels_file_returns_empty(tmp_path: Path) -> None:
    """Свежий однолистовой workbook без чартов/гиперссылок не несёт
    xl/worksheets/_rels/sheet1.xml.rels вовсе."""
    raw = tmp_path / "raw.xlsx"
    Workbook().save(raw)
    with zipfile.ZipFile(raw) as z:
        assert _rel_targets(z, "xl/worksheets/sheet1.xml") == {}


def test_resolve_target_absolute_path_strips_leading_slash() -> None:
    assert _resolve_target("xl/worksheets/sheet1.xml", "/xl/media/image1.png") == "xl/media/image1.png"


def test_resolve_target_relative_path_joins_with_source_dir() -> None:
    """openpyxl сам всегда пишет АБСОЛЮТНЫЕ targets (живая трассировка подтвердила:
    '/xl/worksheets/sheet1.xml' и т.п.) — эта ветка не бьётся реальными openpyxl-фикстурами,
    но легитимна по OPC-спеке (другие писатели/руками собранный OOXML используют
    относительные targets) — прямой unit-тест на чистой функции."""
    assert _resolve_target("xl/worksheets/sheet1.xml", "../drawings/drawing1.xml") == "xl/drawings/drawing1.xml"


# --- _sheet_parts (чистая, синтетический in-memory zip без openpyxl — сурово битые
# workbook.xml, которые ни один реальный писатель не породит, но парсер обязан пережить) ---


def test_sheet_parts_missing_workbook_xml_returns_empty() -> None:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w"):
        pass
    buf.seek(0)
    with zipfile.ZipFile(buf) as z:
        assert _sheet_parts(z) == {}


def test_sheet_parts_skips_sheet_element_without_rid() -> None:
    wb_root = etree.Element(_q("main", "workbook"))
    sheets_el = etree.SubElement(wb_root, _q("main", "sheets"))
    etree.SubElement(sheets_el, _q("main", "sheet")).set("name", "Data")  # без r:id
    wb_xml = etree.tostring(wb_root, xml_declaration=True, encoding="UTF-8", standalone=True)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("xl/workbook.xml", wb_xml)
    buf.seek(0)
    with zipfile.ZipFile(buf) as z:
        assert _sheet_parts(z) == {}


# --- extract_charts: малформед/недостижимая ссылка на каждом шаге OOXML-цепочки
# (терминальная защита от битого файла, докстрока extract_charts) — реальный openpyxl-вывод
# с целенаправленно удалённой частью цепочки, не искусственно упрощённый XML ---


def _remove_zip_part(raw: Path, part_name: str) -> None:
    """In-place: пересобрать zip БЕЗ указанной части — симулирует битую/недостижимую ссылку
    в OOXML-цепочке на реальном openpyxl-выводе."""
    orig = raw.read_bytes()
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(orig)) as z, zipfile.ZipFile(buf, "w") as zo:
        for n in z.namelist():
            if n != part_name:
                zo.writestr(n, z.read(n))
    raw.write_bytes(buf.getvalue())


def _find_part(raw: Path, *, startswith: str, endswith: str = ".xml") -> str:
    with zipfile.ZipFile(raw) as z:
        return next(n for n in z.namelist() if n.startswith(startswith) and n.endswith(endswith))


def test_extract_charts_skips_when_sheet_rels_missing(tmp_path: Path) -> None:
    """drawing r:id есть в самом листе, но .rels листа целиком отсутствует -> drid not in {}."""
    raw = _workbook_with_chart(tmp_path)
    sheet_rels = _find_part(raw, startswith="xl/worksheets/_rels/", endswith=".rels")
    _remove_zip_part(raw, sheet_rels)
    assert extract_charts(raw) == []


def test_extract_charts_skips_when_drawing_part_missing(tmp_path: Path) -> None:
    """Rels листа корректно ссылается на drawing, но сам xl/drawings/drawingN.xml удалён."""
    raw = _workbook_with_chart(tmp_path)
    drawing_part = _find_part(raw, startswith="xl/drawings/", endswith=".xml")
    _remove_zip_part(raw, drawing_part)
    assert extract_charts(raw) == []


def test_extract_charts_skips_when_drawing_rels_missing(tmp_path: Path) -> None:
    """Drawing ссылается на chart-парт, но .rels самого drawing удалён целиком."""
    raw = _workbook_with_chart(tmp_path)
    drawing_rels = _find_part(raw, startswith="xl/drawings/_rels/", endswith=".rels")
    _remove_zip_part(raw, drawing_rels)
    assert extract_charts(raw) == []


def test_extract_charts_skips_when_chart_part_missing(tmp_path: Path) -> None:
    """Drawing rels корректно ссылается на chart, но сам xl/charts/chartN.xml удалён."""
    raw = _workbook_with_chart(tmp_path)
    chart_part = _find_part(raw, startswith="xl/charts/", endswith=".xml")
    _remove_zip_part(raw, chart_part)
    assert extract_charts(raw) == []


# --- _chart_refs (чистая функция — вручную собранное XML, без openpyxl) ---


def test_chart_refs_skips_formula_without_sheet_qualifier() -> None:
    root = etree.Element(_q("c", "chartSpace"))
    etree.SubElement(root, _q("c", "f")).text = "A1"  # без "!" — не наш формат
    assert _chart_refs(root) == []


def test_extract_charts_sheet_with_no_drawing_returns_no_charts(tmp_path: Path) -> None:
    wb = Workbook()
    ws = _active(wb)
    ws.append(["plain", "data"])
    raw = tmp_path / "raw.xlsx"
    wb.save(raw)
    assert extract_charts(raw) == []
