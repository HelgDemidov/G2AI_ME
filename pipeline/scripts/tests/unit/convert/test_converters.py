"""Тесты реестра конвертеров: resolve_converter, детекция скана (_detect_scan)."""
from __future__ import annotations

import datetime as _dt
import hashlib
import logging
import subprocess
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from convert.converters import (
    DOCX_IMAGE_MIN_BYTES,
    ConversionError,
    NeedsOCR,
    UnsupportedFormat,
    _cached_or_call_cloud,
    _capture_original_sha256,
    _check_langs_available,
    cloud_allowed,
    _convert_docx,
    _convert_html,
    _convert_pdf,
    _convert_xlsx,
    _detect_scan,
    _docx_image_markers,
    _docx_referenced_media_ids,
    _ocr_normalize,
    _tesseract_langs,
    _was_ocr_normalized,
    resolve_converter,
    was_ocr_normalized,
)
from core import fsio, schema
from core.schema import SourceRecord
from tests.support import (
    build_docx_with_choice_only_images,
    build_docx_with_group_and_standalone_image,
    build_docx_with_inline_chart,
    build_docx_with_inline_chart_data,
    build_docx_with_inline_image,
    build_docx_with_shape_group,
    build_minimal_docx,
    valid_record,
)


def test_resolve_converter_pdf(tmp_path: Path) -> None:
    conv = resolve_converter(tmp_path / "raw.pdf")
    assert conv.name == "pdf"


def test_resolve_converter_unsupported_lists_known_formats(tmp_path: Path) -> None:
    with pytest.raises(UnsupportedFormat, match="pdf"):
        resolve_converter(tmp_path / "raw.xyz")


def test_resolve_converter_uppercase_extension(tmp_path: Path) -> None:
    conv = resolve_converter(tmp_path / "raw.PDF")
    assert conv.name == "pdf"


def test_resolve_converter_html(tmp_path: Path) -> None:
    conv = resolve_converter(tmp_path / "raw.html")
    assert conv.name == "html"


def test_resolve_converter_docx(tmp_path: Path) -> None:
    conv = resolve_converter(tmp_path / "raw.docx")
    assert conv.name == "docx"


def test_resolve_converter_xlsx(tmp_path: Path) -> None:
    conv = resolve_converter(tmp_path / "raw.xlsx")
    assert conv.name == "xlsx"


# --- _convert_html: реальная trafilatura на инлайн-фикстуре (ни сети, ни модели) ---

_HTML_FIXTURE = """<!doctype html>
<html><head><title>Test Act</title></head>
<body>
<nav><a href="/">Home</a><a href="/about">About</a></nav>
<header>Site chrome — should not survive extraction</header>
<article>
<h1>Article 1</h1>
<p>This is the operative text of the article, long enough to be recognised as
real content by trafilatura's boilerplate-vs-content heuristics rather than
being pruned as a short unimportant fragment.</p>
<table>
<tr><th>Term</th><th>Definition</th></tr>
<tr><td>AI system</td><td>A machine-based system.</td></tr>
</table>
</article>
<footer>Copyright footer — should not survive extraction</footer>
</body></html>"""


def test_convert_html_extracts_article_and_table(tmp_path: Path) -> None:
    raw = tmp_path / "raw.html"
    raw.write_text(_HTML_FIXTURE, encoding="utf-8")
    out = tmp_path / "out.md"
    _convert_html(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    assert "Article 1" in text
    assert "operative text of the article" in text
    assert "AI system" in text and "machine-based system" in text  # таблица сохранена
    assert "Site chrome" not in text  # nav — не контент
    assert "Copyright footer" not in text


def test_html_heading_loss_guard_warns(monkeypatch: Any, tmp_path: Path, caplog: Any) -> None:
    """B2: исходный HTML несёт <h2>, но trafilatura (замокана) отдала выход без
    единого markdown-заголовка — генерическая ловушка «главный контент без
    <article> теряет <hN>» (эмпирика convert-html); НЕ отказ, только warning."""
    raw = tmp_path / "raw.html"
    raw.write_text(
        "<html><body><h2>Some Heading</h2><p>Body text long enough to survive.</p></body></html>",
        encoding="utf-8",
    )
    out = tmp_path / "out.md"
    monkeypatch.setattr("trafilatura.extract", lambda *a, **kw: "Some Heading\n\nBody text long enough to survive.")
    with caplog.at_level(logging.WARNING):
        _convert_html(raw, out, "en")
    assert "вероятна потеря структуры" in caplog.text


def test_html_heading_preserved_no_warning(monkeypatch: Any, tmp_path: Path, caplog: Any) -> None:
    raw = tmp_path / "raw.html"
    raw.write_text("<html><body><h2>Some Heading</h2><p>Body text.</p></body></html>", encoding="utf-8")
    out = tmp_path / "out.md"
    monkeypatch.setattr("trafilatura.extract", lambda *a, **kw: "## Some Heading\n\nBody text.")
    with caplog.at_level(logging.WARNING):
        _convert_html(raw, out, "en")
    assert caplog.text == ""


def test_convert_html_empty_content_raises(tmp_path: Path) -> None:
    """С favor_recall=True даже голая nav-строка выживет как fallback-контент (боязнь
    потерь > боязнь шума, см. spec §design rationale) — по-настоящему пустой ConversionError
    ловит случай, где extract() вернул None целиком (напр. пустой <body>)."""
    raw = tmp_path / "raw.html"
    raw.write_text("<html><body></body></html>", encoding="utf-8")
    out = tmp_path / "out.md"
    with pytest.raises(ConversionError, match="не извлекла"):
        _convert_html(raw, out, "en")


# --- _convert_docx: mammoth+markdownify напрямую на минимальном OOXML (spec
# convert-docx §2-bis.3) — ни сети, ни бинарников в git: фикстура рождается в
# тесте заново ---


def test_convert_docx_extracts_paragraph_text(tmp_path: Path) -> None:
    raw = tmp_path / "raw.docx"
    raw.write_bytes(build_minimal_docx(["Prvi pasus dokumenta.", "Drugi pasus, sa detaljima."]))
    out = tmp_path / "out.md"
    _convert_docx(raw, out, "cnr")
    text = out.read_text(encoding="utf-8")
    assert "Prvi pasus dokumenta." in text
    assert "Drugi pasus, sa detaljima." in text


def test_convert_docx_empty_document_raises(tmp_path: Path) -> None:
    raw = tmp_path / "raw.docx"
    raw.write_bytes(build_minimal_docx([]))
    out = tmp_path / "out.md"
    with pytest.raises(ConversionError, match="mammoth/markdownify не извлекли"):
        _convert_docx(raw, out, "en")


def test_convert_docx_whitespace_only_document_raises(tmp_path: Path) -> None:
    """Параграфы есть, но текст — сплошные пробелы: .strip() должен считать
    это пустым результатом, а не «успешной» конвертацией в пустой файл."""
    raw = tmp_path / "raw.docx"
    raw.write_bytes(build_minimal_docx(["   ", "\t"]))
    out = tmp_path / "out.md"
    with pytest.raises(ConversionError, match="mammoth/markdownify не извлекли"):
        _convert_docx(raw, out, "en")


# --- _docx_image_markers / §2-bis.3: маркеры растров под figures-VLM, только
# для media, реально референсированных документом (orphan-фильтр) ---


def test_docx_image_markers_no_media_at_all(tmp_path: Path) -> None:
    raw = tmp_path / "raw.docx"
    raw.write_bytes(build_minimal_docx(["Body, no images."]))
    assert _docx_image_markers(raw) == ""


def test_convert_docx_no_figures_section_without_large_images(tmp_path: Path) -> None:
    raw = tmp_path / "raw.docx"
    raw.write_bytes(build_minimal_docx(["Main body text."]))
    out = tmp_path / "out.md"
    _convert_docx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    assert "## Figures" not in text


def test_convert_docx_inline_image_marker_positioned(tmp_path: Path) -> None:
    raw = tmp_path / "raw.docx"
    big = b"x" * (DOCX_IMAGE_MIN_BYTES + 1)
    raw.write_bytes(build_docx_with_inline_image(["Before."], big, ["After."]))
    out = tmp_path / "out.md"
    _convert_docx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    id12 = hashlib.sha256(big).hexdigest()[:12]
    assert text.find("Before.") < text.find(id12) < text.find("After.")
    assert "## Figures" not in text


def test_convert_docx_choice_only_image_falls_back_to_figures(tmp_path: Path) -> None:
    raw = tmp_path / "raw.docx"
    big = b"x" * (DOCX_IMAGE_MIN_BYTES + 1)
    raw.write_bytes(build_docx_with_choice_only_images(["Body."], {"c.png": big}))
    out = tmp_path / "out.md"
    _convert_docx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    id12 = hashlib.sha256(big).hexdigest()[:12]
    assert "## Figures (position unknown)" in text
    assert id12 in text


def test_convert_docx_orphan_media_gets_no_marker(tmp_path: Path) -> None:
    """build_minimal_docx(media=...) кладёt файл в word/media/ БЕЗ единой
    ссылки в document.xml.rels — сирота, orphan-фильтр не даёт ей маркера
    (живой кейс: тестовая вырезка отчёта, 21/28 больших media — сироты)."""
    raw = tmp_path / "raw.docx"
    big = b"x" * (DOCX_IMAGE_MIN_BYTES + 1)
    raw.write_bytes(build_minimal_docx(["Main body text."], media={"chart.png": big}))
    out = tmp_path / "out.md"
    _convert_docx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    id12 = hashlib.sha256(big).hexdigest()[:12]
    assert "## Figures" not in text
    assert id12 not in text


def test_docx_referenced_media_ids_wired_vs_orphan(tmp_path: Path) -> None:
    big = b"x" * (DOCX_IMAGE_MIN_BYTES + 1)
    id12 = hashlib.sha256(big).hexdigest()[:12]

    wired = tmp_path / "wired.docx"
    wired.write_bytes(build_docx_with_inline_image(["Before."], big, ["After."]))
    assert id12 in _docx_referenced_media_ids(wired)

    orphan = tmp_path / "orphan.docx"
    orphan.write_bytes(build_minimal_docx(["Body."], media={"chart.png": big}))
    assert _docx_referenced_media_ids(orphan) == frozenset()


def test_docx_image_markers_orphan_silent(tmp_path: Path) -> None:
    raw = tmp_path / "raw.docx"
    big = b"x" * (DOCX_IMAGE_MIN_BYTES + 1)
    raw.write_bytes(build_minimal_docx(["Body."], media={"chart.png": big}))
    assert _docx_image_markers(raw) == ""


def test_docx_image_markers_duplicate_bytes_single_marker(tmp_path: Path) -> None:
    """Два файла с одинаковыми байтами (Word иногда сохраняет один логотип дважды
    под разными именами), оба референсированы -> ОДИН маркер (дедуп по id12,
    в отличие от v1, где каждое вхождение давало свою строку)."""
    raw = tmp_path / "raw.docx"
    same = b"y" * (DOCX_IMAGE_MIN_BYTES + 100)
    raw.write_bytes(build_docx_with_choice_only_images(["Body."], {"a.png": same, "b.png": same}))
    markers = _docx_image_markers(raw)
    expected_id = hashlib.sha256(same).hexdigest()[:12]
    assert markers.count(expected_id) == 1


def test_docx_image_markers_exact_threshold_referenced(tmp_path: Path) -> None:
    """Ровно порог (len == DOCX_IMAGE_MIN_BYTES) не должен считаться маленьким
    (строгое '<', тот же принцип, что SCAN_MIN_TEXTPAGE_FRACTION)."""
    raw = tmp_path / "raw.docx"
    exact = b"x" * DOCX_IMAGE_MIN_BYTES
    raw.write_bytes(build_docx_with_choice_only_images(["Body."], {"chart.png": exact}))
    assert _docx_image_markers(raw) != ""


def test_docx_image_markers_placed_excluded(tmp_path: Path) -> None:
    raw = tmp_path / "raw.docx"
    big = b"x" * (DOCX_IMAGE_MIN_BYTES + 1)
    raw.write_bytes(build_docx_with_choice_only_images(["Body."], {"c.png": big}))
    id12 = hashlib.sha256(big).hexdigest()[:12]
    assert _docx_image_markers(raw, placed=frozenset({id12})) == ""
    assert id12 in _docx_image_markers(raw)


# --- _convert_docx / composite-группы (spec convert-docx §2-ter): Word рисует
# сложную инфографику группой фигур (mc:AlternateContent/wpg:wgp) — вырезается
# ЦЕЛИКОМ pre-проходом docx_groups ДО mammoth, заменяется маркером с captions ---


def test_convert_docx_group_marker_positioned_with_captions(tmp_path: Path) -> None:
    raw = tmp_path / "raw.docx"
    big = b"x" * (DOCX_IMAGE_MIN_BYTES + 1)
    raw.write_bytes(
        build_docx_with_shape_group(["Before."], ["Caption A", "Caption B"], {"a.png": big}, ["After."])
    )
    out = tmp_path / "out.md"
    _convert_docx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    assert "[Figure, docx group " in text
    assert "> captions: Caption A; Caption B" in text
    assert text.find("Before.") < text.find("[Figure, docx group") < text.find("After.")
    assert "DOCXGROUPSENTINEL" not in text


def test_convert_docx_chart_marker_positioned_with_title(tmp_path: Path) -> None:
    """Нативный c:chart (kind="chart") БЕЗ numCache (``build_docx_with_inline_chart``
    несёт только ``c:title`` — до расширения §2-ter mammoth терял чарт МОЛЧА,
    ни маркера, ни текста): caption-фолбэк (chart-data-extraction §4.2) даёт
    ТОТ ЖЕ маркер, что и до data-driven рефакторинга — на месте чарта в потоке.
    Data-driven путь с реальным numCache — см.
    ``test_convert_docx_chart_with_numcache_renders_data_driven_block``."""
    raw = tmp_path / "raw.docx"
    raw.write_bytes(build_docx_with_inline_chart(["Before."], ["Costs of LTE and 5G"], ["After."]))
    out = tmp_path / "out.md"
    _convert_docx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    assert "[Figure, docx chart " in text
    assert "chart content not analyzed" in text
    assert "> captions: Costs of LTE and 5G" in text
    assert text.find("Before.") < text.find("[Figure, docx chart") < text.find("After.")
    assert "DOCXGROUPSENTINEL" not in text


def test_convert_docx_chart_with_numcache_renders_data_driven_block(tmp_path: Path) -> None:
    """Живой факт (chart-data-extraction spec §4.2): нативный c:chart с
    реальным numCache получает ПОЛНЫЙ data-driven рендер (mermaid+таблица,
    отформатированная по value_format) IN-PLACE сентинела — позиция в потоке
    (§4.4: докс-провенанс = сама позиция) сохранена точно, отдельной строки
    провенанса, в отличие от xlsx, не требуется."""
    raw = tmp_path / "raw.docx"
    raw.write_bytes(
        build_docx_with_inline_chart_data(
            ["Before."], ["After."], title="Regional Scores",
            categories=["Montenegro", "Estonia"], values=["0.42", "0.87"], value_format="0.0%",
        )
    )
    out = tmp_path / "out.md"
    _convert_docx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    assert "[Figure, docx chart" not in text
    assert "chart content not analyzed" not in text
    assert "```mermaid\nxychart-beta" in text
    assert "| Montenegro | 42.0% |" in text
    assert text.find("Before.") < text.find("Regional Scores") < text.find("After.")
    assert "DOCXGROUPSENTINEL" not in text


def test_convert_docx_group_media_absorbed_no_individual_markers(tmp_path: Path) -> None:
    """Картинки ВНУТРИ группы не должны всплыть ни инлайн (mammoth их не видит
    — поддерево вырезано ДО неё), ни во фолбэк-секции (поглощены через
    placed=all_media_ids(groups)). id12 маркера группы — хэш XML-поддерева
    (см. docx_groups.extract_and_strip_groups), НЕ хэш байт картинки — байтовый
    id картинки нигде в выводе не печатается, используется только внутри для
    исключения из фолбэка."""
    raw = tmp_path / "raw.docx"
    big = b"x" * (DOCX_IMAGE_MIN_BYTES + 1)
    raw.write_bytes(build_docx_with_shape_group(["Before."], ["Cap"], {"a.png": big}, ["After."]))
    out = tmp_path / "out.md"
    _convert_docx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    image_id12 = hashlib.sha256(big).hexdigest()[:12]
    assert "docx media" not in text  # ни одного индивидуального image-маркера
    assert "## Figures" not in text  # и во фолбэк не улетело
    assert image_id12 not in text  # байтовый id картинки нигде не всплывает
    assert "[Figure, docx group " in text  # групповой маркер (свой id12) есть


def test_convert_docx_no_groups_behaves_like_v2(tmp_path: Path) -> None:
    """Документ без composite-групп — поведение v2 не изменилось (регресс-guard)."""
    raw = tmp_path / "raw.docx"
    big = b"x" * (DOCX_IMAGE_MIN_BYTES + 1)
    raw.write_bytes(build_docx_with_inline_image(["Before."], big, ["After."]))
    out = tmp_path / "out.md"
    _convert_docx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    id12 = hashlib.sha256(big).hexdigest()[:12]
    assert f"> [Image, docx media {id12} — raster content not analyzed]" in text
    assert "docx group" not in text


def test_convert_docx_group_alongside_standalone_image(tmp_path: Path) -> None:
    """Смешанный документ: composite-группа + ОДИНОЧНАЯ картинка вне группы —
    оба пути (§2-bis инлайн + §2-ter групповой маркер) работают одновременно,
    без перекрёстного заражения (группа не проглатывает чужую картинку,
    одиночная картинка не всплывает как ещё один групповой маркер)."""
    raw = tmp_path / "raw.docx"
    group_img = b"g" * (DOCX_IMAGE_MIN_BYTES + 1)
    standalone_img = b"s" * (DOCX_IMAGE_MIN_BYTES + 1)
    raw.write_bytes(build_docx_with_group_and_standalone_image(["Cap"], group_img, standalone_img))
    out = tmp_path / "out.md"
    _convert_docx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    group_image_id12 = hashlib.sha256(group_img).hexdigest()[:12]
    standalone_id = hashlib.sha256(standalone_img).hexdigest()[:12]
    assert "[Figure, docx group " in text  # групповой маркер (свой id12 = хэш AC-поддерева)
    assert f"> [Image, docx media {standalone_id} — raster content not analyzed]" in text
    assert "## Figures" not in text  # ни один из двух не улетел во фолбэк
    assert group_image_id12 not in text  # байтовый id картинки ГРУППЫ нигде не всплывает
    assert text.count(standalone_id) == 1


# --- _convert_xlsx: openpyxl.Workbook() in-memory фикстуры (spec convert-xlsx
# §2, тестовое покрытие) — ни сети, ни бинарников в git ---


def _active(wb: Any) -> Any:
    """``wb.active`` типизирован как ``Worksheet | None`` в стабах — свежий
    ``Workbook()`` всегда несёт активный лист, узкий helper вместо ``assert``
    в каждом тесте."""
    ws = wb.active
    assert ws is not None
    return ws


def _save_wb(tmp_path: Path, wb: Any, name: str = "raw.xlsx") -> Path:
    raw = tmp_path / name
    wb.save(raw)
    return raw


def _xlsx_with_cached_formula(tmp_path: Path, header: str, formula: str, cached: str) -> Path:
    """Ручной патч сохранённого openpyxl-файла: вписывает закэшированное
    значение формулы (``<v>``) — сам openpyxl формулы не считает, а живой
    xlsx несёт кэш от Excel/LibreOffice (data_only=True его и читает)."""
    import io
    import zipfile

    wb = openpyxl.Workbook()
    ws = _active(wb)
    ws.append([header])
    ws.append([formula])
    buf = io.BytesIO()
    wb.save(buf)
    with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as z:
        names = z.namelist()
        sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    patched = sheet_xml.replace("<v></v>", f"<v>{cached}</v>")
    raw = tmp_path / "raw.xlsx"
    with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as z, zipfile.ZipFile(raw, "w") as zo:
        for n in names:
            zo.writestr(n, patched if n == "xl/worksheets/sheet1.xml" else z.read(n))
    return raw


def test_convert_xlsx_single_sheet_simple_values_preserve_order(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = _active(wb)
    ws.title = "Data"
    ws.append(["Name", "Count"])
    ws.append(["Alpha", 1])
    ws.append(["Beta", 2])
    raw = _save_wb(tmp_path, wb)
    out = tmp_path / "out.md"
    _convert_xlsx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    assert "## Data" in text
    assert "| Name | Count |" in text
    assert "| --- | --- |" in text
    assert "| Alpha | 1 |" in text
    assert "| Beta | 2 |" in text
    assert text.index("| Name | Count |") < text.index("| Alpha | 1 |") < text.index("| Beta | 2 |")


def test_convert_xlsx_merged_cell_value_only_in_anchor(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = _active(wb)
    ws.append(["Header1", "Header2"])
    ws.append(["Merged Value", None])
    ws.merge_cells("A2:B2")
    raw = _save_wb(tmp_path, wb)
    out = tmp_path / "out.md"
    _convert_xlsx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    assert "| Merged Value |  |" in text  # соседняя ячейка диапазона пуста


def test_convert_xlsx_hidden_sheet_marked_but_content_kept(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws1 = _active(wb)
    ws1.title = "Visible"
    ws1.append(["A"])
    ws2 = wb.create_sheet("Calc")
    ws2.append(["secret helper value"])
    ws2.sheet_state = "hidden"
    raw = _save_wb(tmp_path, wb)
    out = tmp_path / "out.md"
    _convert_xlsx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    assert "## Calc (hidden)" in text
    assert "secret helper value" in text  # честность: контент не теряется, не orphan-фильтр


def test_convert_xlsx_empty_sheet_gets_marker_not_empty_table(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws1 = _active(wb)
    ws1.title = "Data"
    ws1.append(["x"])
    wb.create_sheet("Blank")
    raw = _save_wb(tmp_path, wb)
    out = tmp_path / "out.md"
    _convert_xlsx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    assert '## Blank\n\n> [Sheet "Blank" — empty, skipped]' in text
    assert "| --- |" not in text.split('> [Sheet "Blank"')[1]  # ни одной таблицы после маркера


def test_convert_xlsx_date_cell_renders_iso(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = _active(wb)
    ws.append(["Date"])
    ws.append([_dt.date(2026, 7, 20)])
    raw = _save_wb(tmp_path, wb)
    out = tmp_path / "out.md"
    _convert_xlsx(raw, out, "en")
    assert "2026-07-20" in out.read_text(encoding="utf-8")


def test_convert_xlsx_formula_with_cached_value_used_not_formula_text(tmp_path: Path) -> None:
    raw = _xlsx_with_cached_formula(tmp_path, "Header", "=1+1", "4")
    out = tmp_path / "out.md"
    _convert_xlsx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    assert "| 4 |" in text
    assert "1+1" not in text


def test_convert_xlsx_formula_without_cache_renders_empty_not_crash(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = _active(wb)
    ws.append(["Header"])
    ws.append(["=1+1"])
    raw = _save_wb(tmp_path, wb)
    out = tmp_path / "out.md"
    _convert_xlsx(raw, out, "en")  # не должно упасть на None-значении формулы без кэша
    text = out.read_text(encoding="utf-8")
    assert "| Header |" in text


def test_convert_xlsx_all_sheets_empty_raises(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()  # дефолтный лист без единой записанной ячейки
    raw = _save_wb(tmp_path, wb)
    out = tmp_path / "out.md"
    with pytest.raises(ConversionError):
        _convert_xlsx(raw, out, "en")


# --- _convert_xlsx / встроенные чарты (spec convert-xlsx §3): маркер сразу
# после таблицы своего листа, chart-only лист не теряется как «пустой» ---


def test_convert_xlsx_chart_marker_positioned_after_sheet_table(tmp_path: Path) -> None:
    """``openpyxl.chart``'s writer (``add_data``) never populates ``c:numCache``
    (empirically confirmed, chart-data-extraction spec) — this chart has no
    cached data, so ``_render_xlsx_chart_block`` falls back to the honest
    caption marker (data-driven table+mermaid path is covered separately, see
    ``test_convert_xlsx_chart_with_numcache_renders_data_driven_block``)."""
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()
    ws = _active(wb)
    ws.title = "Data"
    ws.append(["Cat", "Val"])
    ws.append(["A", 1])
    chart = BarChart()
    chart.title = "My Chart"
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=2), titles_from_data=True)
    ws.add_chart(chart, "D2")
    raw = _save_wb(tmp_path, wb)
    out = tmp_path / "out.md"
    _convert_xlsx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    assert "| Cat | Val |" in text
    assert "> [Figure, xlsx chart " in text
    assert "on Data!D2 — chart content not analyzed" in text
    assert "> captions: My Chart" in text
    assert text.index("| A | 1 |") < text.index("[Figure, xlsx chart")


def test_convert_xlsx_chart_only_sheet_not_marked_empty(tmp_path: Path) -> None:
    """Лист без единой заполненной ячейки, но с висящим на нём чартом —
    честный chart-маркер (fallback — openpyxl не пишет numCache, см. тест
    выше), НЕ «[Sheet ... — empty, skipped]» (иначе чарт молча теряется)."""
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()
    ws1 = _active(wb)
    ws1.title = "Data"
    ws1.append(["Cat", "Val"])
    ws1.append(["A", 1])
    ws2 = wb.create_sheet("ChartOnly")
    chart = BarChart()
    chart.title = "Orphan Chart"
    chart.add_data(Reference(ws1, min_col=2, min_row=1, max_row=2), titles_from_data=True)
    ws2.add_chart(chart, "A1")
    raw = _save_wb(tmp_path, wb)
    out = tmp_path / "out.md"
    _convert_xlsx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    assert '> [Sheet "ChartOnly" — empty, skipped]' not in text
    assert "on ChartOnly!A1" in text


def _xlsx_with_chart_numcache(
    tmp_path: Path,
    *,
    sheet_name: str = "Data",
    anchor: str = "D2",
    title: str = "My Chart",
    categories: list[str] | None = None,
    values: list[str] | None = None,
    value_format: str = "0.0%",
) -> Path:
    """openpyxl-сгенерированный workbook с чартом (корректная структура
    worksheet/drawing/rels) -> ручной патч ``xl/charts/chart1.xml`` РЕАЛЬНЫМ
    ``c:numCache``/``c:strCache`` (тот же приём, что ``_xlsx_with_cached_formula``
    выше: openpyxl не пишет кэш, живой Excel/LibreOffice — пишет; см. spec
    chart-data-extraction §1)."""
    import io
    import zipfile

    from openpyxl.chart import BarChart, Reference

    cats, vals = categories or ["A", "B"], values or ["1", "2"]
    wb = openpyxl.Workbook()
    ws = _active(wb)
    ws.title = sheet_name
    ws.append(["Cat", "Val"])
    for c, v in zip(cats, vals, strict=True):
        ws.append([c, v])
    chart = BarChart()
    chart.title = title
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=len(cats) + 1), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(cats) + 1))
    ws.add_chart(chart, anchor)
    buf = io.BytesIO()
    wb.save(buf)

    cat_pts = "".join(f'<c:pt idx="{i}"><c:v>{c}</c:v></c:pt>' for i, c in enumerate(cats))
    val_pts = "".join(f'<c:pt idx="{i}"><c:v>{v}</c:v></c:pt>' for i, v in enumerate(vals))
    chart_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<c:chartSpace xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f'<c:chart><c:title><c:tx><c:rich><a:p><a:r><a:t>{title}</a:t></a:r></a:p></c:rich></c:tx></c:title>'
        '<c:plotArea><c:barChart><c:barDir val="col"/><c:grouping val="clustered"/>'
        '<c:ser><c:idx val="0"/><c:order val="0"/>'
        f'<c:tx><c:strRef><c:f>{sheet_name}!$B$1</c:f><c:strCache><c:ptCount val="1"/>'
        '<c:pt idx="0"><c:v>Val</c:v></c:pt></c:strCache></c:strRef></c:tx>'
        f'<c:cat><c:strRef><c:f>{sheet_name}!$A$2:$A${len(cats) + 1}</c:f>'
        f'<c:strCache><c:ptCount val="{len(cats)}"/>{cat_pts}</c:strCache></c:strRef></c:cat>'
        f'<c:val><c:numRef><c:f>{sheet_name}!$B$2:$B${len(vals) + 1}</c:f>'
        f'<c:numCache><c:formatCode>{value_format}</c:formatCode>'
        f'<c:ptCount val="{len(vals)}"/>{val_pts}</c:numCache></c:numRef></c:val>'
        '</c:ser></c:barChart></c:plotArea></c:chart></c:chartSpace>'
    )

    with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as z:
        names = z.namelist()
        contents = {n: z.read(n) for n in names}
    contents["xl/charts/chart1.xml"] = chart_xml.encode()
    raw = tmp_path / "raw.xlsx"
    with zipfile.ZipFile(raw, "w") as zo:
        for n in names:
            zo.writestr(n, contents[n])
    return raw


def test_convert_xlsx_chart_with_numcache_renders_data_driven_block(tmp_path: Path) -> None:
    """Живой факт (chart-data-extraction spec §1): в отличие от openpyxl-
    записанного чарта (fallback-тесты выше), чарт с реальным numCache
    получает ПОЛНЫЙ data-driven рендер: провенанс-строка (§4.4) + mermaid +
    отформатированная по value_format таблица, НЕ старый VLM-маркер."""
    raw = _xlsx_with_chart_numcache(
        tmp_path, sheet_name="Data", anchor="D2", title="My Chart",
        categories=["A", "B"], values=["0.42", "0.87"], value_format="0.0%",
    )
    out = tmp_path / "out.md"
    _convert_xlsx(raw, out, "en")
    text = out.read_text(encoding="utf-8")
    assert "> лист Data, якорь D2" in text
    assert "```mermaid\nxychart-beta" in text
    assert "| Category | Val |" in text
    assert "| A | 42.0% |" in text
    assert "| B | 87.0% |" in text
    assert "[Figure, xlsx chart" not in text
    assert text.index("| Cat | Val |") < text.index("> лист Data") < text.index("```mermaid")


# --- _tesseract_langs: rec.language -> tesseract -l аргумент ---


def test_tesseract_langs_english() -> None:
    assert _tesseract_langs("en") == "eng"


def test_tesseract_langs_latin_script_gets_eng_suffix() -> None:
    assert _tesseract_langs("cnr") == "srp_latn+eng"
    assert _tesseract_langs("et") == "est+eng"
    assert _tesseract_langs("es") == "spa+eng"


def test_tesseract_langs_cjk_and_arabic_no_eng_suffix() -> None:
    assert _tesseract_langs("zh") == "chi_sim"
    assert _tesseract_langs("ja") == "jpn"
    assert _tesseract_langs("ar") == "ara"


def test_tesseract_langs_unknown_code_falls_back_to_eng(caplog: Any) -> None:
    result = _tesseract_langs("xx")
    assert result == "eng"
    assert "xx" in caplog.text


def test_tesseract_langs_none_defaults_to_eng_silently(caplog: Any) -> None:
    result = _tesseract_langs(None)
    assert result == "eng"
    assert caplog.text == ""


# --- _detect_scan: fake pdfplumber (паттерн test_pdf_to_markdown._FakeEmptyPdf) ---


class _FakePage:
    def __init__(self, text: str) -> None:
        self._text = text

    def extract_text(self) -> str | None:
        return self._text


class _FakePdf:
    def __init__(self, pages: list[_FakePage], metadata: dict[str, str] | None = None) -> None:
        self.pages = pages
        self.metadata = metadata or {}

    def __enter__(self) -> "_FakePdf":
        return self

    def __exit__(self, *exc: Any) -> None:
        return None


def _patch_open(monkeypatch: Any, pages: list[Any], metadata: dict[str, str] | None = None) -> None:
    monkeypatch.setattr(
        "convert.converters.pdfplumber.open", lambda path: _FakePdf(pages, metadata)
    )


def test_detect_scan_raises_when_all_pages_empty(monkeypatch: Any, tmp_path: Path) -> None:
    _patch_open(monkeypatch, [_FakePage(""), _FakePage(""), _FakePage("")])
    with pytest.raises(NeedsOCR, match="0/3"):
        _detect_scan(tmp_path / "raw.pdf")


def test_detect_scan_passes_when_half_pages_have_text(monkeypatch: Any, tmp_path: Path) -> None:
    """Ровно порог (with_text/n == 0.5) не должен считаться сканом (строгое '<')."""
    long_text = "x" * 60  # >= SCAN_MIN_CHARS_PER_PAGE
    _patch_open(monkeypatch, [_FakePage(long_text), _FakePage("")])
    _detect_scan(tmp_path / "raw.pdf")  # не бросает


def test_detect_scan_noop_on_zero_pages(monkeypatch: Any, tmp_path: Path) -> None:
    _patch_open(monkeypatch, [])
    _detect_scan(tmp_path / "raw.pdf")  # диагностирует pdf_convert, не _detect_scan


def test_detect_scan_none_extract_text_treated_as_empty(monkeypatch: Any, tmp_path: Path) -> None:
    class _NonePage:
        def extract_text(self) -> str | None:
            return None

    _patch_open(monkeypatch, [_NonePage(), _NonePage()])
    with pytest.raises(NeedsOCR):
        _detect_scan(tmp_path / "raw.pdf")


# --- _was_ocr_normalized: метка ocrmypdf в метаданных переживает мутацию raw ---


def test_was_ocr_normalized_true_when_creator_mentions_ocrmypdf(monkeypatch: Any, tmp_path: Path) -> None:
    _patch_open(monkeypatch, [], metadata={"Creator": "ocrmypdf 15.2.0+dfsg1 / Tesseract OCR-PDF 5.3.4"})
    assert _was_ocr_normalized(tmp_path / "raw.pdf") is True


def test_was_ocr_normalized_false_for_born_digital_pdf(monkeypatch: Any, tmp_path: Path) -> None:
    _patch_open(monkeypatch, [], metadata={"Creator": "Microsoft® Word 2019"})
    assert _was_ocr_normalized(tmp_path / "raw.pdf") is False


def test_was_ocr_normalized_false_when_no_metadata_at_all(monkeypatch: Any, tmp_path: Path) -> None:
    _patch_open(monkeypatch, [])  # metadata=None -> {}
    assert _was_ocr_normalized(tmp_path / "raw.pdf") is False


def test_public_was_ocr_normalized_alias_matches_private(monkeypatch: Any, tmp_path: Path) -> None:
    """Публичный алиас (discovery-snowball §2.3: потребитель вне convert-слоя) должен
    давать ИДЕНТИЧНЫЙ результат приватной реализации на одном и том же входе — иначе
    рефактор одной из двух функций разошёлся бы незамеченным."""
    _patch_open(monkeypatch, [], metadata={"Creator": "ocrmypdf 15.2.0"})
    raw = tmp_path / "raw.pdf"
    assert was_ocr_normalized(raw) is True
    assert was_ocr_normalized(raw) == _was_ocr_normalized(raw)


def test_convert_pdf_calls_pdf_convert_only_after_scan_check(monkeypatch: Any, tmp_path: Path) -> None:
    calls: list[str] = []
    _patch_open(monkeypatch, [_FakePage("x" * 60)])
    monkeypatch.setattr(
        "convert.converters.pdf_convert",
        lambda src, dst: calls.append("pdf_convert"),
    )
    _convert_pdf(tmp_path / "raw.pdf", tmp_path / "out.md", "en")
    assert calls == ["pdf_convert"]


def test_convert_pdf_routes_scan_through_ocr_normalize(monkeypatch: Any, tmp_path: Path) -> None:
    """С convert-ocr NeedsOCR больше не пропагируется наружу — скан идёт через
    _ocr_normalize (мутирует raw IN-PLACE, один файл — не сайдкар), pdf_convert
    затем вызывается на ТОМ ЖЕ raw, и вывод проходит post-проход ocr_headings
    (только OCR-ветка)."""
    _patch_open(monkeypatch, [_FakePage(""), _FakePage("")])
    monkeypatch.setattr("convert.converters.cloud_allowed", lambda record: False)  # локальный путь явно
    raw = tmp_path / "raw.pdf"
    raw.write_bytes(b"fake scanned pdf")

    normalize_calls: list[tuple[Path, str | None]] = []

    def fake_normalize(raw_arg: Path, language: str | None) -> None:
        normalize_calls.append((raw_arg, language))

    monkeypatch.setattr("convert.converters._ocr_normalize", fake_normalize)
    convert_calls: list[tuple[str, str]] = []

    def fake_pdf_convert(src: str, dst: str) -> None:
        convert_calls.append((src, dst))
        Path(dst).write_text("ANNEX I\nSome body text.\n", encoding="utf-8")

    monkeypatch.setattr("convert.converters.pdf_convert", fake_pdf_convert)
    out = tmp_path / "out.md"
    _convert_pdf(raw, out, "en")
    assert normalize_calls == [(raw, "en")]
    assert convert_calls == [(str(raw), str(out))]  # тот же raw, не отдельный файл
    assert out.read_text(encoding="utf-8") == "# ANNEX I\nSome body text.\n"  # ocr_headings применён


def test_convert_pdf_digital_path_skips_ocr_headings_postprocessing(
    monkeypatch: Any, tmp_path: Path
) -> None:
    """Цифровой путь (текст-слой есть) НЕ проходит ocr_headings — во избежание
    регресса калиброванной размер-кластеризации pdf_to_markdown (Сингапур/Эстония)."""
    _patch_open(monkeypatch, [_FakePage("x" * 60)])  # текст есть — не скан

    def fake_pdf_convert(src: str, dst: str) -> None:
        Path(dst).write_text("ANNEX I\nSome body text.\n", encoding="utf-8")

    monkeypatch.setattr("convert.converters.pdf_convert", fake_pdf_convert)
    out = tmp_path / "out.md"
    _convert_pdf(tmp_path / "raw.pdf", out, "en")
    assert out.read_text(encoding="utf-8") == "ANNEX I\nSome body text.\n"  # без изменений


def test_convert_pdf_reapplies_ocr_headings_on_already_normalized_raw(
    monkeypatch: Any, tmp_path: Path
) -> None:
    """Регресс: raw уже нормализован ПРЕДЫДУЩИМ прогоном (текст-слой есть, _detect_scan
    больше не поднимет NeedsOCR) — но ocr_headings обязан примениться СНОВА на этой
    повторной конвертации (--force/бамп версии конвертера), иначе восстановление
    заголовков теряется после первого же прогона. Метка — метаданные ocrmypdf,
    переживающие мутацию текст-слоя (см. _was_ocr_normalized)."""
    _patch_open(
        monkeypatch, [_FakePage("x" * 60)],
        metadata={"Creator": "ocrmypdf 15.2.0+dfsg1 / Tesseract OCR-PDF 5.3.4"},
    )
    monkeypatch.setattr("convert.converters.cloud_allowed", lambda record: False)  # локальный путь явно

    def fake_pdf_convert(src: str, dst: str) -> None:
        Path(dst).write_text("ANNEX I\nSome body text.\n", encoding="utf-8")

    monkeypatch.setattr("convert.converters.pdf_convert", fake_pdf_convert)
    out = tmp_path / "out.md"
    _convert_pdf(tmp_path / "raw.pdf", out, "en")
    assert out.read_text(encoding="utf-8") == "# ANNEX I\nSome body text.\n"  # ocr_headings применён


# --- _check_langs_available / _ocr_normalize ---


def test_check_langs_available_all_present(monkeypatch: Any) -> None:
    fake = subprocess.CompletedProcess(
        args=[], returncode=0,
        stdout="List of available languages (3):\neng\nsrp_latn\nest\n", stderr="",
    )
    monkeypatch.setattr("convert.converters.subprocess.run", lambda *a, **kw: fake)
    _check_langs_available("srp_latn+eng")  # не бросает


def test_check_langs_available_missing_raises_with_apt_command(monkeypatch: Any) -> None:
    fake = subprocess.CompletedProcess(
        args=[], returncode=0, stdout="List of available languages (1):\neng\n", stderr=""
    )
    monkeypatch.setattr("convert.converters.subprocess.run", lambda *a, **kw: fake)
    with pytest.raises(ConversionError, match="tesseract-ocr-chi-sim"):
        _check_langs_available("chi_sim")


def test_ocr_normalize_missing_binary_raises_needs_ocr_with_apt_command(
    monkeypatch: Any, tmp_path: Path
) -> None:
    raw = tmp_path / "raw.pdf"
    raw.write_bytes(b"%PDF-1.4 fake")
    monkeypatch.setattr("convert.converters.shutil.which", lambda name: None)
    with pytest.raises(NeedsOCR, match="apt install ocrmypdf"):
        _ocr_normalize(raw, "en")


def test_ocr_normalize_mutates_raw_in_place_no_sidecar_file(monkeypatch: Any, tmp_path: Path) -> None:
    """Один PDF-файл на документ — не сайдкар .ocr.pdf (решение пользователя, ревизия
    convert-ocr от 2026-07-17): raw.pdf заменяется версией с текст-слоем НА МЕСТЕ."""
    raw = tmp_path / "raw.pdf"
    raw.write_bytes(b"%PDF-1.4 fake scan")

    monkeypatch.setattr("convert.converters.shutil.which", lambda name: "/usr/bin/ocrmypdf")
    monkeypatch.setattr("convert.converters._check_langs_available", lambda langs: None)
    _patch_open(monkeypatch, [_FakePage("")])

    calls: list[list[str]] = []

    def fake_run(cmd: list[str], **kw: Any) -> Any:
        calls.append(cmd)
        Path(cmd[-1]).write_bytes(b"%PDF-1.4 ocr result")  # ocrmypdf пишет staging
        return subprocess.CompletedProcess(args=cmd, returncode=0, stdout="", stderr="")

    monkeypatch.setattr("convert.converters.subprocess.run", fake_run)
    _ocr_normalize(raw, "en")
    assert calls  # subprocess реально вызван
    assert raw.read_bytes() == b"%PDF-1.4 ocr result"  # raw САМ теперь содержит OCR-результат
    assert not (tmp_path / ".ocr.pdf").exists()  # никакого сайдкара


def test_ocr_normalize_always_calls_subprocess_no_cache_check(monkeypatch: Any, tmp_path: Path) -> None:
    """Без сайдкара нечего и проверять на свежесть — _ocr_normalize безусловно
    вызывает subprocess при каждом вызове (кэш получается «бесплатно» иначе: после
    успеха raw САМ содержит текст, и _detect_scan больше не поднимет NeedsOCR — то
    есть _ocr_normalize вообще не будет вызван на следующих прогонах, без кэш-файла)."""
    raw = tmp_path / "raw.pdf"
    raw.write_bytes(b"%PDF-1.4 fake scan")
    monkeypatch.setattr("convert.converters.shutil.which", lambda name: "/usr/bin/ocrmypdf")
    monkeypatch.setattr("convert.converters._check_langs_available", lambda langs: None)
    _patch_open(monkeypatch, [_FakePage("")])

    calls: list[list[str]] = []

    def fake_run(cmd: list[str], **kw: Any) -> Any:
        calls.append(cmd)
        Path(cmd[-1]).write_bytes(b"ocr result")
        return subprocess.CompletedProcess(args=cmd, returncode=0, stdout="", stderr="")

    monkeypatch.setattr("convert.converters.subprocess.run", fake_run)
    _ocr_normalize(raw, "en")
    assert len(calls) == 1


def test_ocr_normalize_nonzero_exit_raises_conversion_error_with_stderr_tail(
    monkeypatch: Any, tmp_path: Path
) -> None:
    raw = tmp_path / "raw.pdf"
    raw.write_bytes(b"%PDF-1.4 fake")
    monkeypatch.setattr("convert.converters.shutil.which", lambda name: "/usr/bin/ocrmypdf")
    monkeypatch.setattr("convert.converters._check_langs_available", lambda langs: None)
    _patch_open(monkeypatch, [_FakePage("")])

    def fake_run(cmd: list[str], **kw: Any) -> Any:
        return subprocess.CompletedProcess(args=cmd, returncode=1, stdout="", stderr="boom: bad scan")

    monkeypatch.setattr("convert.converters.subprocess.run", fake_run)
    with pytest.raises(ConversionError, match="boom: bad scan"):
        _ocr_normalize(raw, "en")


def test_ocr_normalize_warns_on_large_page_count(
    monkeypatch: Any, tmp_path: Path, caplog: Any
) -> None:
    raw = tmp_path / "raw.pdf"
    raw.write_bytes(b"%PDF-1.4 fake")
    monkeypatch.setattr("convert.converters.shutil.which", lambda name: "/usr/bin/ocrmypdf")
    monkeypatch.setattr("convert.converters._check_langs_available", lambda langs: None)
    _patch_open(monkeypatch, [_FakePage("")] * 250)  # > OCR_PAGE_WARN

    def fake_run(cmd: list[str], **kw: Any) -> Any:
        Path(cmd[-1]).write_bytes(b"ocr result")
        return subprocess.CompletedProcess(args=cmd, returncode=0, stdout="", stderr="")

    monkeypatch.setattr("convert.converters.subprocess.run", fake_run)
    with caplog.at_level(logging.WARNING):
        _ocr_normalize(raw, "en")
    assert "250" in caplog.text


# --- _capture_original_sha256 (spec ocr-eval-harness §8.1, S1) ---


def test_capture_original_sha256_stores_hash_of_current_file(tmp_path: Path) -> None:
    raw = tmp_path / "raw.pdf"
    raw.write_bytes(b"%PDF-1.4 pristine scan bytes")
    expected = fsio.sha256_file(raw)

    _capture_original_sha256(raw)

    state = schema.load_state(raw.parent / ".state.yaml")
    assert state.original_sha256 == expected


def test_capture_original_sha256_does_not_overwrite_existing_value(tmp_path: Path) -> None:
    """Второй прогон (--force/бамп версии) на УЖЕ мутированном raw не должен
    затереть ранее захваченный оригинальный хэш пересчитанным от мутированного
    файла — иначе original_sha256 перестаёт означать «издательский оригинал»."""
    raw = tmp_path / "raw.pdf"
    raw.write_bytes(b"first content")
    state_path = raw.parent / ".state.yaml"
    schema.save_state(state_path, schema.OperationalState(original_sha256="f" * 64))

    raw.write_bytes(b"different content entirely")  # симулирует уже-мутированный файл
    _capture_original_sha256(raw)

    assert schema.load_state(state_path).original_sha256 == "f" * 64  # не тронуто


def test_ocr_normalize_captures_hash_before_ocrmypdf_mutates_raw(monkeypatch: Any, tmp_path: Path) -> None:
    """Полный поток: захваченный original_sha256 — хэш ИСХОДНЫХ байт raw, не
    результата ocrmypdf (мок subprocess.run переписывает raw другим содержимым)."""
    raw = tmp_path / "raw.pdf"
    original_bytes = b"%PDF-1.4 pristine scan"
    raw.write_bytes(original_bytes)
    expected = fsio.sha256_file(raw)

    monkeypatch.setattr("convert.converters.shutil.which", lambda name: "/usr/bin/ocrmypdf")
    monkeypatch.setattr("convert.converters._check_langs_available", lambda langs: None)
    _patch_open(monkeypatch, [_FakePage("")])

    def fake_run(cmd: list[str], **kw: Any) -> Any:
        Path(cmd[-1]).write_bytes(b"totally different ocr-mutated bytes")
        return subprocess.CompletedProcess(args=cmd, returncode=0, stdout="", stderr="")

    monkeypatch.setattr("convert.converters.subprocess.run", fake_run)
    _ocr_normalize(raw, "en")

    assert raw.read_bytes() != original_bytes  # raw реально мутирован
    state = schema.load_state(raw.parent / ".state.yaml")
    assert state.original_sha256 == expected  # но захвачен хэш ДО мутации


def test_ocr_normalize_second_call_keeps_first_captured_hash(monkeypatch: Any, tmp_path: Path) -> None:
    """Регресс на --force повторную нормализацию: второй вызов _ocr_normalize
    (уже на мутированном raw) не должен переписать original_sha256."""
    raw = tmp_path / "raw.pdf"
    raw.write_bytes(b"pristine")
    expected = fsio.sha256_file(raw)

    monkeypatch.setattr("convert.converters.shutil.which", lambda name: "/usr/bin/ocrmypdf")
    monkeypatch.setattr("convert.converters._check_langs_available", lambda langs: None)
    _patch_open(monkeypatch, [_FakePage("")])

    def fake_run_v1(cmd: list[str], **kw: Any) -> Any:
        Path(cmd[-1]).write_bytes(b"ocr v1")
        return subprocess.CompletedProcess(args=cmd, returncode=0, stdout="", stderr="")

    monkeypatch.setattr("convert.converters.subprocess.run", fake_run_v1)
    _ocr_normalize(raw, "en")  # первый прогон — захватывает expected

    def fake_run_v2(cmd: list[str], **kw: Any) -> Any:
        Path(cmd[-1]).write_bytes(b"ocr v2")
        return subprocess.CompletedProcess(args=cmd, returncode=0, stdout="", stderr="")

    monkeypatch.setattr("convert.converters.subprocess.run", fake_run_v2)
    _ocr_normalize(raw, "en")  # второй прогон (--force) — raw уже не оригинал

    state = schema.load_state(raw.parent / ".state.yaml")
    assert state.original_sha256 == expected  # осталось от первого прогона


def test_convert_pdf_digital_path_leaves_original_sha256_unset(monkeypatch: Any, tmp_path: Path) -> None:
    """Born-digital документ никогда не проходит через _ocr_normalize ->
    original_sha256 остаётся None — полю нечего означать без OCR-мутации."""
    _patch_open(monkeypatch, [_FakePage("x" * 60)])  # текст есть — не скан
    raw = tmp_path / "raw.pdf"
    raw.write_bytes(b"born digital pdf bytes")

    def fake_pdf_convert(src: str, dst: str) -> None:
        Path(dst).write_text("Body text.\n", encoding="utf-8")

    monkeypatch.setattr("convert.converters.pdf_convert", fake_pdf_convert)
    _convert_pdf(raw, tmp_path / "out.md", "en")

    state = schema.load_state(raw.parent / ".state.yaml")
    assert state.original_sha256 is None


# --- cloud_allowed / _cached_or_call_cloud / _convert_pdf облачная маршрутизация
# (spec convert-cloud-tier §6) ---


def _record(**over: Any) -> SourceRecord:
    data = valid_record()
    data.update(over)
    return SourceRecord.model_validate(data)


def _reset_cloud_module_state(monkeypatch: Any) -> None:
    monkeypatch.setattr("convert.converters._CLOUD_DISABLED", False)
    monkeypatch.setattr("convert.converters._CLOUD_KEY_WARNED", False)


def test_cloud_allowed_false_when_disabled_flag_set(monkeypatch: Any) -> None:
    _reset_cloud_module_state(monkeypatch)
    monkeypatch.setattr("convert.converters._CLOUD_DISABLED", True)
    monkeypatch.setenv("OPENROUTER_API_KEY", "test-key")
    assert cloud_allowed(None) is False


def test_cloud_allowed_false_for_confidential_record(monkeypatch: Any) -> None:
    _reset_cloud_module_state(monkeypatch)
    monkeypatch.setenv("OPENROUTER_API_KEY", "test-key")
    rec = _record(sensitivity="confidential")
    assert cloud_allowed(rec) is False


def test_cloud_allowed_false_without_key(monkeypatch: Any) -> None:
    _reset_cloud_module_state(monkeypatch)
    monkeypatch.delenv("OPENROUTER_API_KEY", raising=False)
    monkeypatch.setattr("convert.converters.load_dotenv", lambda: None)  # не подхватить реальный .env
    assert cloud_allowed(None) is False


def test_cloud_allowed_true_for_normal_record_with_key(monkeypatch: Any) -> None:
    _reset_cloud_module_state(monkeypatch)
    monkeypatch.setenv("OPENROUTER_API_KEY", "test-key")
    rec = _record(sensitivity="normal")
    assert cloud_allowed(rec) is True


def test_cached_or_call_cloud_hit_skips_network(monkeypatch: Any, tmp_path: Path) -> None:
    raw = tmp_path / "raw.pdf"
    raw.write_bytes(b"normalized scan bytes")
    from core import fsio, schema

    cache = raw.parent / ".cloudocr.md"
    cache.write_text("# Cached\n\nBody.", encoding="utf-8")
    state = schema.OperationalState(cloud_ocr_model="m", cloud_ocr_raw_sha256=fsio.sha256_file(raw))
    schema.save_state(raw.parent / ".state.yaml", state)

    monkeypatch.setattr(
        "convert.converters.cloud_ocr.convert_scan",
        lambda *a, **kw: (_ for _ in ()).throw(AssertionError("сеть не должна была вызываться")),
    )
    text = _cached_or_call_cloud(raw, "en", model="m")
    assert text == "# Cached\n\nBody."


def test_cached_or_call_cloud_model_mismatch_uses_cache_without_recall(
    monkeypatch: Any, tmp_path: Path, caplog: Any
) -> None:
    raw = tmp_path / "raw.pdf"
    raw.write_bytes(b"normalized scan bytes")
    from core import fsio, schema

    cache = raw.parent / ".cloudocr.md"
    cache.write_text("# From old model", encoding="utf-8")
    state = schema.OperationalState(cloud_ocr_model="old-model", cloud_ocr_raw_sha256=fsio.sha256_file(raw))
    schema.save_state(raw.parent / ".state.yaml", state)

    monkeypatch.setattr(
        "convert.converters.cloud_ocr.convert_scan",
        lambda *a, **kw: (_ for _ in ()).throw(AssertionError("авто-перевызов при смене модели запрещён")),
    )
    with caplog.at_level(logging.WARNING):
        text = _cached_or_call_cloud(raw, "en", model="new-model")
    assert text == "# From old model"
    assert "old-model" in caplog.text and "new-model" in caplog.text


def test_cached_or_call_cloud_miss_calls_cloud_and_persists(monkeypatch: Any, tmp_path: Path) -> None:
    raw = tmp_path / "raw.pdf"
    raw.write_bytes(b"normalized scan bytes")
    from core import fsio, schema

    monkeypatch.setattr("convert.converters.cloud_ocr.convert_scan", lambda raw_, lang, *, model: "# Fresh\n\nText.")
    text = _cached_or_call_cloud(raw, "en", model="m")
    assert text == "# Fresh\n\nText."
    assert (raw.parent / ".cloudocr.md").read_text(encoding="utf-8") == "# Fresh\n\nText."
    state = schema.load_state(raw.parent / ".state.yaml")
    assert state.cloud_ocr_model == "m"
    assert state.cloud_ocr_raw_sha256 == fsio.sha256_file(raw)


def test_cached_or_call_cloud_failure_returns_none(monkeypatch: Any, tmp_path: Path, caplog: Any) -> None:
    raw = tmp_path / "raw.pdf"
    raw.write_bytes(b"normalized scan bytes")

    def failing(*a: Any, **kw: Any) -> str:
        raise RuntimeError("OpenRouter: исчерпаны попытки")

    monkeypatch.setattr("convert.converters.cloud_ocr.convert_scan", failing)
    with caplog.at_level(logging.WARNING):
        text = _cached_or_call_cloud(raw, "en", model="m")
    assert text is None
    assert not (raw.parent / ".cloudocr.md").exists()


def test_convert_pdf_confidential_record_skips_cloud(monkeypatch: Any, tmp_path: Path) -> None:
    _reset_cloud_module_state(monkeypatch)
    monkeypatch.setenv("OPENROUTER_API_KEY", "test-key")
    _patch_open(monkeypatch, [_FakePage("x" * 60)], metadata={"Creator": "ocrmypdf 15.2.0"})  # уже нормализован

    monkeypatch.setattr(
        "convert.converters._cached_or_call_cloud",
        lambda *a, **kw: (_ for _ in ()).throw(AssertionError("облако не должно было вызываться")),
    )

    def fake_pdf_convert(src: str, dst: str) -> None:
        Path(dst).write_text("ANNEX I\nBody.\n", encoding="utf-8")

    monkeypatch.setattr("convert.converters.pdf_convert", fake_pdf_convert)
    out = tmp_path / "out.md"
    rec = _record(sensitivity="confidential")
    _convert_pdf(tmp_path / "raw.pdf", out, "en", record=rec)
    assert out.read_text(encoding="utf-8") == "# ANNEX I\nBody.\n"  # локальный путь + ocr_headings


def test_convert_pdf_digital_never_calls_cloud(monkeypatch: Any, tmp_path: Path) -> None:
    """Цифровой PDF (не скан) — облако не вызывается вовсе, независимо от гейтов."""
    _reset_cloud_module_state(monkeypatch)
    monkeypatch.setenv("OPENROUTER_API_KEY", "test-key")
    _patch_open(monkeypatch, [_FakePage("x" * 60)])  # текст есть — не скан

    monkeypatch.setattr(
        "convert.converters._cached_or_call_cloud",
        lambda *a, **kw: (_ for _ in ()).throw(AssertionError("облако не должно было вызываться")),
    )

    def fake_pdf_convert(src: str, dst: str) -> None:
        Path(dst).write_text("body", encoding="utf-8")

    monkeypatch.setattr("convert.converters.pdf_convert", fake_pdf_convert)
    out = tmp_path / "out.md"
    _convert_pdf(tmp_path / "raw.pdf", out, "en", record=_record())
    assert out.read_text(encoding="utf-8") == "body"


def test_convert_pdf_cloud_success_applies_additive_merge_not_full_promote(
    monkeypatch: Any, tmp_path: Path
) -> None:
    """Облачный вывод проходит ТОЛЬКО additive-режим (§2.5, v2.1): пропущенная облаком
    CAPS-глава промоутится в ##, а существующая облачная разметка неприкосновенна —
    «# Cloud Title» полный promote_flat_headings СНЯЛ бы (не-CAPS строка), additive
    обязан оставить. Живой мотиватор: главы I.–VIII. me-crps (чекпоинт 1)."""
    _reset_cloud_module_state(monkeypatch)
    monkeypatch.setenv("OPENROUTER_API_KEY", "test-key")
    _patch_open(monkeypatch, [_FakePage("x" * 60)], metadata={"Creator": "ocrmypdf 15.2.0"})

    monkeypatch.setattr(
        "convert.converters._cached_or_call_cloud",
        lambda raw, lang, *, model: "# Cloud Title\n\nII. GLAVNI DIO TEKSTA\n\nBody, unflagged.",
    )
    monkeypatch.setattr(
        "convert.converters.pdf_convert",
        lambda src, dst: (_ for _ in ()).throw(AssertionError("локальный путь не должен был вызываться")),
    )
    out = tmp_path / "out.md"
    _convert_pdf(tmp_path / "raw.pdf", out, "en", record=_record())
    assert out.read_text(encoding="utf-8") == (
        "# Cloud Title\n\n## II. GLAVNI DIO TEKSTA\n\nBody, unflagged."
    )


def test_convert_pdf_cloud_failure_falls_back_to_local_with_ocr_headings(
    monkeypatch: Any, tmp_path: Path
) -> None:
    _reset_cloud_module_state(monkeypatch)
    monkeypatch.setenv("OPENROUTER_API_KEY", "test-key")
    _patch_open(monkeypatch, [_FakePage("x" * 60)], metadata={"Creator": "ocrmypdf 15.2.0"})

    monkeypatch.setattr("convert.converters._cached_or_call_cloud", lambda raw, lang, *, model: None)

    def fake_pdf_convert(src: str, dst: str) -> None:
        Path(dst).write_text("ANNEX I\nBody.\n", encoding="utf-8")

    monkeypatch.setattr("convert.converters.pdf_convert", fake_pdf_convert)
    out = tmp_path / "out.md"
    _convert_pdf(tmp_path / "raw.pdf", out, "en", record=_record())
    assert out.read_text(encoding="utf-8") == "# ANNEX I\nBody.\n"  # локальный фолбэк + ocr_headings, без краха
