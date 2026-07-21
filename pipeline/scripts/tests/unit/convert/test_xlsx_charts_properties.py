"""Property-based тесты интервальной арифметики xlsx_charts.py (test-coverage-hardening,
раунд 2 — Hypothesis, общепроектный стандарт с pdf_to_markdown-раунда, §3.B-bis.2).

Существующие example-тесты покрывают конкретные вручную собранные сценарии (в т.ч. живой
CE12-дефект — data_keep vs exclude на реальной мини-книге); эти тесты фаззят те же инварианты
на случайных диапазонах — модуль-докстрока extract_chart_workbook описывает ТРИ раунда живых
дефектов именно в этой геометрии, прямое лекарство от того же класса риска, что у
detect_columns в pdf_to_markdown."""
from __future__ import annotations

from typing import Any

from hypothesis import given, settings
from hypothesis import strategies as st
from lxml import etree
from openpyxl.utils import get_column_letter

from convert.xlsx_charts import (
    _blank_foreign_cells,
    _bounding_union,
    _is_compact,
    _pad_range_axes,
    _q,
    _range_contains,
)

_MAX_COORD = 200


@st.composite
def _ranges(draw: Any, max_coord: int = _MAX_COORD) -> tuple[int, int, int, int]:
    """Случайный (col_lo, row_lo, col_hi, row_hi), 1-indexed, col_lo<=col_hi, row_lo<=row_hi —
    та же форма, что реальные диапазоны данных/bbox в xlsx_charts.py."""
    c1 = draw(st.integers(min_value=1, max_value=max_coord))
    c2 = draw(st.integers(min_value=c1, max_value=max_coord))
    r1 = draw(st.integers(min_value=1, max_value=max_coord))
    r2 = draw(st.integers(min_value=r1, max_value=max_coord))
    return (c1, r1, c2, r2)


# --- _pad_range_axes: результат ⊇ входа, клэмп на 1 ---


@given(
    rng=_ranges(),
    col_pad=st.integers(min_value=0, max_value=20),
    row_pad=st.integers(min_value=0, max_value=20),
)
@settings(max_examples=100)
def test_pad_range_axes_result_contains_original(
    rng: tuple[int, int, int, int], col_pad: int, row_pad: int
) -> None:
    c1, r1, c2, r2 = rng
    pc1, pr1, pc2, pr2 = _pad_range_axes(rng, col_pad, row_pad)
    assert pc1 <= c1 and pr1 <= r1 and pc2 >= c2 and pr2 >= r2


@given(
    rng=_ranges(),
    col_pad=st.integers(min_value=0, max_value=500),
    row_pad=st.integers(min_value=0, max_value=500),
)
@settings(max_examples=100)
def test_pad_range_axes_never_below_one(
    rng: tuple[int, int, int, int], col_pad: int, row_pad: int
) -> None:
    pc1, pr1, _pc2, _pr2 = _pad_range_axes(rng, col_pad, row_pad)
    assert pc1 >= 1
    assert pr1 >= 1


# --- _range_contains: собственные углы всегда внутри; монотонность под запасом ---


@given(rng=_ranges())
@settings(max_examples=100)
def test_range_contains_own_corners(rng: tuple[int, int, int, int]) -> None:
    c1, r1, c2, r2 = rng
    assert _range_contains(rng, c1, r1)
    assert _range_contains(rng, c2, r2)


@given(
    rng=_ranges(),
    col_pad=st.integers(min_value=0, max_value=20),
    row_pad=st.integers(min_value=0, max_value=20),
    col=st.integers(min_value=1, max_value=_MAX_COORD),
    row=st.integers(min_value=1, max_value=_MAX_COORD),
)
@settings(max_examples=100)
def test_range_contains_monotonic_under_padding(
    rng: tuple[int, int, int, int], col_pad: int, row_pad: int, col: int, row: int
) -> None:
    """Точка, попадающая в исходный диапазон, попадает и в его же диапазон с ЛЮБЫМ
    неотрицательным запасом — запас может только расширять, никогда не сужать."""
    if _range_contains(rng, col, row):
        padded = _pad_range_axes(rng, col_pad, row_pad)
        assert _range_contains(padded, col, row)


# --- _bounding_union: объемлет все входы, границы — истинные min/max ---


@given(ranges=st.lists(_ranges(), min_size=1, max_size=10))
@settings(max_examples=100)
def test_bounding_union_contains_all_corners(ranges: list[tuple[int, int, int, int]]) -> None:
    union = _bounding_union(ranges)
    for rng in ranges:
        c1, r1, c2, r2 = rng
        assert _range_contains(union, c1, r1)
        assert _range_contains(union, c2, r2)


@given(ranges=st.lists(_ranges(), min_size=1, max_size=10))
@settings(max_examples=100)
def test_bounding_union_is_true_min_max(ranges: list[tuple[int, int, int, int]]) -> None:
    union_c1, union_r1, union_c2, union_r2 = _bounding_union(ranges)
    assert union_c1 == min(r[0] for r in ranges)
    assert union_r1 == min(r[1] for r in ranges)
    assert union_c2 == max(r[2] for r in ranges)
    assert union_r2 == max(r[3] for r in ranges)


# --- _is_compact: согласована с прямым вычислением ширины/высоты ---


@given(rng=_ranges(), max_span=st.integers(min_value=1, max_value=_MAX_COORD))
@settings(max_examples=100)
def test_is_compact_matches_direct_computation(rng: tuple[int, int, int, int], max_span: int) -> None:
    c1, r1, c2, r2 = rng
    expected = (c2 - c1 + 1) <= max_span and (r2 - r1 + 1) <= max_span
    assert _is_compact(rng, max_span) == expected


# --- _blank_foreign_cells: ГЛАВНЫЙ инвариант — data_keep ВСЕГДА побеждает exclude
# (прямой фаззинг класса дефекта «CE12», см. докстроку extract_chart_workbook §3) ---


def _single_cell_sheet(col: int, row: int) -> Any:
    root = etree.Element(_q("main", "worksheet"))
    sheet_data = etree.SubElement(root, _q("main", "sheetData"))
    row_el = etree.SubElement(sheet_data, _q("main", "row"))
    row_el.set("r", str(row))
    etree.SubElement(row_el, _q("main", "c")).set("r", f"{get_column_letter(col)}{row}")
    return root


def _surviving_refs(sheet_root: Any) -> set[str]:
    sheet_data = sheet_root.find(_q("main", "sheetData"))
    return {c.get("r") for row_el in sheet_data for c in row_el}


@st.composite
def _range_containing(draw: Any, col: int, row: int, max_coord: int = _MAX_COORD) -> tuple[int, int, int, int]:
    """Диапазон, ГАРАНТИРОВАННО содержащий (col, row) — строится вокруг точки, не полагается
    на assume()-фильтрацию (та начала отбрасывать >85% случайных независимых диапазонов)."""
    c1 = draw(st.integers(min_value=1, max_value=col))
    c2 = draw(st.integers(min_value=col, max_value=max_coord))
    r1 = draw(st.integers(min_value=1, max_value=row))
    r2 = draw(st.integers(min_value=row, max_value=max_coord))
    return (c1, r1, c2, r2)


@given(
    col=st.integers(min_value=1, max_value=100),
    row=st.integers(min_value=1, max_value=100),
    extra_data_keep=st.lists(_ranges(max_coord=100), max_size=4),
    exclude=st.lists(_ranges(max_coord=100), max_size=5),
    data=st.data(),
)
@settings(max_examples=200)
def test_blank_foreign_cells_data_keep_always_beats_exclude(
    col: int,
    row: int,
    extra_data_keep: list[tuple[int, int, int, int]],
    exclude: list[tuple[int, int, int, int]],
    data: Any,
) -> None:
    """Ячейка, объявленная СВОЕЙ (data_keep), никогда не зачищается — при каком угодно
    случайно сгенерированном наборе exclude-диапазонов, даже пересекающихся с data_keep.
    Один диапазон data_keep СКОНСТРУИРОВАН так, чтобы гарантированно содержать (col, row)
    (не полагается на assume() — та же health-check ловушка, что у контрольного теста
    ниже); остальные data_keep/exclude — полностью случайны."""
    own_range = data.draw(_range_containing(col, row))
    data_keep = [own_range, *extra_data_keep]
    sheet_root = _single_cell_sheet(col, row)
    _blank_foreign_cells(sheet_root, (1, 1, _MAX_COORD, _MAX_COORD), data_keep, exclude)
    assert f"{get_column_letter(col)}{row}" in _surviving_refs(sheet_root)


@given(
    col=st.integers(min_value=1, max_value=100),
    row=st.integers(min_value=1, max_value=100),
    data=st.data(),
)
@settings(max_examples=100)
def test_blank_foreign_cells_removes_cell_only_in_exclude(col: int, row: int, data: Any) -> None:
    """Контрольный случай: ячейка НЕ в data_keep, но в exclude — вычищается (без этого
    первый тест мог бы пройти тривиально из-за сломанной/no-op реализации)."""
    exclude_range = data.draw(_range_containing(col, row))
    sheet_root = _single_cell_sheet(col, row)
    _blank_foreign_cells(sheet_root, (1, 1, _MAX_COORD, _MAX_COORD), [], [exclude_range])
    assert f"{get_column_letter(col)}{row}" not in _surviving_refs(sheet_root)
